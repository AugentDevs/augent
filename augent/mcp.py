"""
Augent MCP Server

Model Context Protocol server for Claude Code integration.
Exposes Augent as a native tool that Claude can call directly.

Tools exposed:
- download_audio: Download audio from video URLs (YouTube, etc.) at maximum speed
- transcribe_audio: Full transcription without keyword search
- search_audio: Search for keywords in audio files
- deep_search: Semantic search by meaning, not just keywords
- take_notes: All-in-one note-taking: download + transcribe + save .md to Desktop
- chapters: Auto-detect topic chapters in audio
- batch_search: Search multiple audio files in parallel
- text_to_speech: Convert text to natural speech audio using Kokoro TTS
- search_proximity: Find keywords appearing near each other
- identify_speakers: Speaker diarization (who said what)
- list_files: List media files in a directory
- list_memories: List all stored transcriptions
- memory_stats: View transcription memory statistics
- clear_memory: Clear transcription memory
- search_memory: Search across ALL stored transcriptions
- separate_audio: Separate audio into stems (vocals, drums, bass, other) using Demucs v4
- clip_export: Export a video clip from a URL for a specific time range
- highlights: Export the best moments from a transcription as MP4 clips
- tag: Add, remove, or list tags on transcriptions
- rebuild_graph: Rebuild Obsidian graph view data for all transcriptions
- visual: Extract visual context from video at moments that matter
- spaces: Download, check, or stop X/Twitter Spaces recordings

Usage:
  python -m augent.mcp
  # or
  augent-mcp

Add to Claude Code project (.mcp.json):
  {
    "mcpServers": {
      "augent": {
        "command": "augent-mcp"
      }
    }
  }

"""

import json
import os
import signal
import shutil
import subprocess
import sys
import time
import uuid
from typing import Any

# Check for required dependencies before importing
_MISSING_DEPS = []
try:
    import faster_whisper
except ImportError:
    _MISSING_DEPS.append("faster-whisper")

try:
    import torch
except ImportError:
    _MISSING_DEPS.append("torch")

if _MISSING_DEPS:

    def _dependency_error():
        return {
            "error": f"Missing dependencies: {', '.join(_MISSING_DEPS)}. "
            f"Install with: pip install {' '.join(_MISSING_DEPS)}"
        }

    # Create stub functions that return errors
    def search_audio(*args, **kwargs):
        raise RuntimeError(_dependency_error()["error"])

    def search_audio_full(*args, **kwargs):
        raise RuntimeError(_dependency_error()["error"])

    def transcribe_audio(*args, **kwargs):
        raise RuntimeError(_dependency_error()["error"])

    def search_audio_proximity(*args, **kwargs):
        raise RuntimeError(_dependency_error()["error"])

    def get_memory_stats():
        return _dependency_error()

    def clear_memory():
        return 0

    def list_memories():
        return []

else:
    from .core import (
        clear_memory,
        get_memory_stats,
        list_memories,
        search_audio,
        search_audio_full,
        search_audio_proximity,
        transcribe_audio,
    )

# Optional dependencies (sentence-transformers, pyannote-audio, kokoro)
# are imported lazily inside handler functions so that installing them
# mid-session works without restarting the MCP server.


def send_response(response: dict) -> None:
    """Send JSON-RPC response to stdout."""
    output = json.dumps(response)
    sys.stdout.write(output + "\n")
    sys.stdout.flush()


def _strip_quarantine(path: str) -> None:
    """Remove macOS quarantine flag from a file."""
    import platform
    import subprocess

    if platform.system() == "Darwin":
        try:
            subprocess.run(
                ["xattr", "-d", "com.apple.quarantine", path], capture_output=True
            )
        except Exception:
            pass


import re as _re

# Track source URLs for downloaded files so transcription can attach them
_downloaded_urls: dict = {}  # file_path -> source_url

_YOUTUBE_VIDEO_ID_RE = _re.compile(
    r"(?:youtube\.com/watch\?v=|youtu\.be/|youtube\.com/embed/|youtube\.com/shorts/)"
    r"([a-zA-Z0-9_-]{11})"
)


def _extract_youtube_id(url: str) -> str:
    """Extract YouTube video ID from a URL. Returns empty string if not YouTube."""
    if not url:
        return ""
    m = _YOUTUBE_VIDEO_ID_RE.search(url)
    return m.group(1) if m else ""


def _youtube_timestamp_link(source_url: str, seconds: float) -> str:
    """Generate a YouTube URL with timestamp parameter. Returns empty string if not YouTube."""
    video_id = _extract_youtube_id(source_url)
    if not video_id:
        return ""
    return f"https://youtube.com/watch?v={video_id}&t={int(seconds)}"


def _write_output_file(
    output_path: str,
    rows: list,
    columns: list,
    bold_columns: list = None,
    keyword_column: str = None,
) -> str:
    """
    Write results to CSV or XLSX based on file extension.

    Args:
        output_path: File path (.csv or .xlsx)
        rows: List of dicts with data
        columns: Column keys to include
        bold_columns: Column keys to bold in XLSX
        keyword_column: Column key containing text with **bold** keywords
    Returns:
        Absolute path of written file
    """
    import os
    import re

    path = os.path.expanduser(output_path)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    bold_columns = bold_columns or []

    if path.endswith(".xlsx"):
        _write_xlsx(path, rows, columns, bold_columns, keyword_column)
    else:
        _write_csv(path, rows, columns)

    _strip_quarantine(path)
    return os.path.abspath(path)


def _write_csv(path: str, rows: list, columns: list) -> None:
    """Write plain CSV file."""
    import csv
    import io
    import re

    bold_pattern = re.compile(r"\*\*(.+?)\*\*")

    buf = io.StringIO()
    writer = csv.writer(buf)

    # Header
    header_names = {
        "timestamp": "Timestamp",
        "text": "Text",
        "snippet": "Snippet",
        "keyword": "Keyword",
        "timestamp_seconds": "Seconds",
        "confidence": "Confidence",
        "match_type": "Match Type",
        "similarity": "Similarity",
        "source": "Source",
        "title": "Source",
        "youtube_link": "YouTube Link",
    }
    writer.writerow([header_names.get(c, c.title()) for c in columns])

    for row in rows:
        vals = []
        for c in columns:
            v = row.get(c, "")
            if isinstance(v, str):
                v = bold_pattern.sub(r"\1", v)
                v = v.replace("...", "").strip()
            vals.append(v)
        writer.writerow(vals)

    with open(path, "w", newline="") as f:
        f.write(buf.getvalue())


def _write_xlsx(
    path: str, rows: list, columns: list, bold_columns: list, keyword_column: str = None
) -> None:
    """Write styled XLSX file with bold headers, timestamps, and keywords."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        _write_csv(path.replace(".xlsx", ".csv"), rows, columns)
        return

    import re

    bold_pattern = re.compile(r"\*\*(.+?)\*\*")

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Styles
    header_fill = PatternFill(
        start_color="1F1F1F", end_color="1F1F1F", fill_type="solid"
    )
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    thin_border = Border(bottom=Side(style="thin", color="E0E0E0"))

    # Column names
    header_names = {
        "timestamp": "Timestamp",
        "text": "Text",
        "snippet": "Snippet",
        "keyword": "Keyword",
        "timestamp_seconds": "Seconds",
        "confidence": "Confidence",
        "match_type": "Match Type",
        "similarity": "Similarity",
        "source": "Source",
        "title": "Source",
        "youtube_link": "YouTube Link",
    }

    # Write header row
    for col_idx, col_key in enumerate(columns, 1):
        cell = ws.cell(
            row=1, column=col_idx, value=header_names.get(col_key, col_key.title())
        )
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_key in enumerate(columns, 1):
            val = row_data.get(col_key, "")
            if isinstance(val, str):
                val = bold_pattern.sub(r"\1", val)
                val = val.replace("...", "").strip()
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_key in bold_columns:
                cell.font = bold_font
            else:
                cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    # Auto-width columns
    for col_idx, col_key in enumerate(columns, 1):
        max_len = len(header_names.get(col_key, col_key))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
            max_len + 4
        )

    wb.save(path)


def send_error(id: Any, code: int, message: str) -> None:
    """Send JSON-RPC error response."""
    send_response(
        {"jsonrpc": "2.0", "id": id, "error": {"code": code, "message": message}}
    )


def handle_initialize(id: Any, params: dict) -> None:
    """Handle initialize request."""
    send_response(
        {
            "jsonrpc": "2.0",
            "id": id,
            "result": {
                "protocolVersion": "2024-11-05",
                "capabilities": {"tools": {}},
                "serverInfo": {"name": "augent", "version": "2026.3.29"},
            },
        }
    )


_ALL_TOOLS = [
    {
        "name": "download_audio",
        "description": "Download audio from video URLs at maximum speed. Built by Augent with speed optimizations (aria2c multi-connection, concurrent fragments). Downloads audio ONLY - never video. Supports YouTube, Vimeo, TikTok, Twitter, SoundCloud, and 1000+ sites. IMPORTANT: When a user pastes ANY URL and asks what was said, what happened, or wants the content — use this tool to download, then transcribe_audio to get the text. This is the correct path for tweets, videos, podcasts, and any link with audio/video content. Do NOT use WebFetch for media URLs.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "url": {
                    "type": "string",
                    "description": "Video URL to download audio from (YouTube, Vimeo, TikTok, etc.)",
                },
                "output_dir": {
                    "type": "string",
                    "description": "Directory to save the audio file. Default: ~/Downloads",
                },
            },
            "required": ["url"],
        },
    },
    {
        "name": "transcribe_audio",
        "description": "Transcribe an audio file and return the full text with timestamps. Useful when you need the complete transcription rather than searching for specific keywords.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "audio_path": {
                    "type": "string",
                    "description": "Path to the audio file",
                },
                "model_size": {
                    "type": "string",
                    "enum": [
                        "tiny",
                        "base",
                        "small",
                        "medium",
                        "large",
                    ],
                    "description": "Whisper model size. ALWAYS use tiny unless the user explicitly requests a different size. tiny is already highly accurate.",
                },
                "start": {
                    "type": "number",
                    "description": "Start transcription at this many seconds into the audio. Default: 0 (beginning)",
                },
                "duration": {
                    "type": "number",
                    "description": "Only transcribe this many seconds of audio. Example: 600 = first 10 minutes. Default: full file",
                },
                "output": {
                    "type": "string",
                    "description": "Optional file path to save transcription. Use .csv for plain data or .xlsx for styled spreadsheets with bold headers and formatting.",
                },
                "translated_text": {
                    "type": "string",
                    "description": "English translation of a non-English transcription. When provided, no audio processing occurs — the translation is stored alongside the existing cached transcription as a sibling (eng) markdown file. The audio must have been transcribed already. Pass the full English text as a single string.",
                },
            },
            "required": ["audio_path"],
        },
    },
    {
        "name": "search_audio",
        "description": "Search audio files for keywords and return timestamped matches with context snippets. Useful for finding specific moments in podcasts, interviews, lectures, or any audio content.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "audio_path": {
                    "type": "string",
                    "description": "Path to the audio file (MP3, WAV, M4A, etc.)",
                },
                "keywords": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of keywords or phrases to search for",
                },
                "model_size": {
                    "type": "string",
                    "enum": [
                        "tiny",
                        "base",
                        "small",
                        "medium",
                        "large",
                    ],
                    "description": "Whisper model size. ALWAYS use tiny unless the user explicitly requests a different size. tiny is already highly accurate.",
                },
                "include_full_text": {
                    "type": "boolean",
                    "description": "Include full transcription text in response. Default: false",
                },
                "output": {
                    "type": "string",
                    "description": "Optional file path to save results. Use .csv for plain data or .xlsx for styled spreadsheets with bold headers and formatting.",
                },
                "clip": {
                    "type": "boolean",
                    "description": "Download actual video clips around each match. Set to true when the user asks for clips, highlights, compilations, or says they want the video itself, not just timestamps. Requires the audio to have been downloaded from a URL. Default: false",
                },
                "clip_padding": {
                    "type": "integer",
                    "description": "Seconds of padding before and after each match for clip export. Default: 15",
                },
            },
            "required": ["audio_path", "keywords"],
        },
    },
    {
        "name": "deep_search",
        "description": "Search audio by meaning, not just keywords.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "audio_path": {
                    "type": "string",
                    "description": "Path to the audio file",
                },
                "query": {
                    "type": "string",
                    "description": "Natural language search query (e.g. 'discussion about funding challenges')",
                },
                "top_k": {
                    "type": "integer",
                    "description": "Number of results to return. Default: 5",
                },
                "model_size": {
                    "type": "string",
                    "enum": [
                        "tiny",
                        "base",
                        "small",
                        "medium",
                        "large",
                    ],
                    "description": "Whisper model size. ALWAYS use tiny unless the user explicitly requests a different size. tiny is already highly accurate.",
                },
                "output": {
                    "type": "string",
                    "description": "Optional file path to save results. Use .csv for plain data or .xlsx for styled spreadsheets with bold headers and formatting.",
                },
                "context_words": {
                    "type": "integer",
                    "description": "Words of context per result. Default: 25. Use 150 for full evidence blocks when Claude needs to answer a question, not just find a moment.",
                },
                "dedup_seconds": {
                    "type": "number",
                    "description": "Merge matches within this many seconds of each other to avoid redundant results. Default: 0 (off). Use 60 for Q&A.",
                },
                "clip": {
                    "type": "boolean",
                    "description": "Download actual video clips around each match. Set to true when the user asks for clips, highlights, compilations, or says they want the video itself, not just timestamps. Requires the audio to have been downloaded from a URL. Default: false",
                },
                "clip_padding": {
                    "type": "integer",
                    "description": "Seconds of padding before and after each match for clip export. Default: 15",
                },
            },
            "required": ["audio_path", "query"],
        },
    },
    {
        "name": "take_notes",
        "description": "Take notes from a URL. Downloads audio, transcribes, and saves .md to Desktop. This single tool handles the entire pipeline — download, transcribe, and save — when the user asks for notes, summaries, highlights, takeaways, eye-candy, quiz, or any formatted content from a video/audio URL. Returns audio_path for follow-up tools (chapters, search). Also used to SAVE formatted notes: call with save_content to write notes to the file from the previous take_notes call (no url needed).",
        "inputSchema": {
            "type": "object",
            "properties": {
                "url": {
                    "type": "string",
                    "description": "Video/audio URL to take notes from (YouTube, Vimeo, TikTok, Twitter, SoundCloud, etc.)",
                },
                "save_content": {
                    "type": "string",
                    "description": "Formatted notes content to save. When provided, writes this content to a file. Works with a previous take_notes call OR with output_path for saving notes from memory transcripts.",
                },
                "output_path": {
                    "type": "string",
                    "description": "Explicit file path to save notes to. Use this when saving notes from a memory transcript (no prior take_notes url call). E.g. ~/Desktop/My_Notes.md",
                },
                "output_dir": {
                    "type": "string",
                    "description": "Directory to save the .md notes file. Default: ~/Desktop",
                },
                "style": {
                    "type": "string",
                    "enum": [
                        "tldr",
                        "notes",
                        "highlight",
                        "eye-candy",
                        "quiz",
                    ],
                    "description": "Note style. tldr > notes > highlight > eye-candy increases formatting richness. quiz generates questions. Default: notes. Pick based on what the user asks for.",
                },
                "model_size": {
                    "type": "string",
                    "enum": [
                        "tiny",
                        "base",
                        "small",
                        "medium",
                        "large",
                    ],
                    "description": "Whisper model size. ALWAYS use tiny unless the user explicitly requests a different size. tiny is already highly accurate.",
                },
                "read_aloud": {
                    "type": "boolean",
                    "description": "Generate a spoken audio summary and embed it in the notes for Obsidian playback. Default: false",
                },
                "visual": {
                    "type": "string",
                    "description": "Extract visual context from the video. Pass a query describing what needs visual context (e.g. 'the workflow setup steps'). Frames are saved to the Obsidian vault with ![[]] embeds in a Visual Context .md file. Requires video URL (downloads the video automatically).",
                },
            },
        },
    },
    {
        "name": "chapters",
        "description": "Auto-detect topic chapters in audio.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "audio_path": {
                    "type": "string",
                    "description": "Path to the audio file",
                },
                "sensitivity": {
                    "type": "number",
                    "description": "0.0 = many chapters, 1.0 = few chapters. Default: 0.4",
                },
                "model_size": {
                    "type": "string",
                    "enum": [
                        "tiny",
                        "base",
                        "small",
                        "medium",
                        "large",
                    ],
                    "description": "Whisper model size. ALWAYS use tiny unless the user explicitly requests a different size. tiny is already highly accurate.",
                },
            },
            "required": ["audio_path"],
        },
    },
    {
        "name": "batch_search",
        "description": "Search multiple audio files for keywords in parallel. Ideal for processing podcast libraries, interview collections, or any batch of audio files. Returns aggregated results with file paths.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "audio_paths": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of paths to audio files",
                },
                "keywords": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of keywords or phrases to search for",
                },
                "model_size": {
                    "type": "string",
                    "enum": [
                        "tiny",
                        "base",
                        "small",
                        "medium",
                        "large",
                    ],
                    "description": "Whisper model size. ALWAYS use tiny unless the user explicitly requests a different size. tiny is already highly accurate.",
                },
                "workers": {
                    "type": "integer",
                    "description": "Number of parallel workers. Default: 2",
                },
            },
            "required": ["audio_paths", "keywords"],
        },
    },
    {
        "name": "text_to_speech",
        "description": "Convert text to natural speech audio using Kokoro TTS. Saves an MP3 file. Runs in background — returns a job_id immediately. Call again with job_id to check status. Pass text for raw TTS, or file_path to read a notes file (strips markdown, skips metadata, embeds audio player).",
        "inputSchema": {
            "type": "object",
            "properties": {
                "job_id": {
                    "type": "string",
                    "description": "Check status of a running TTS job. Pass the job_id returned from a previous call.",
                },
                "text": {
                    "type": "string",
                    "description": "Text to convert to speech. Either text or file_path is required.",
                },
                "file_path": {
                    "type": "string",
                    "description": "Path to a notes file to read aloud. Strips markdown formatting, skips metadata, generates MP3, and embeds audio player in the file.",
                },
                "voice": {
                    "type": "string",
                    "description": "Voice ID. American English female: af_heart (default), af_alloy, af_aoede, af_bella, af_jessica, af_kore, af_nicole, af_nova, af_river, af_sarah, af_sky. American English male: am_adam, am_echo, am_eric, am_fenrir, am_liam, am_michael, am_onyx, am_puck. British English: bf_emma, bf_isabella, bf_lily, bm_daniel, bm_fable, bm_george, bm_lewis. Other languages: Spanish (ef_dora, em_alex), French (ff_siwis), Hindi (hf_alpha, hf_beta, hm_omega, hm_psi), Italian (if_sara, im_nicola), Japanese (jf_alpha, jf_gongitsune, jf_nezumi, jf_tebukuro, jm_kumo), Brazilian Portuguese (pf_dora, pm_alex), Mandarin Chinese (zf_xiaobei, zf_xiaoni, zf_xiaoxiao, zf_xiaoyi, zm_yunjian, zm_yunxi, zm_yunxia, zm_yunyang).",
                },
                "output_dir": {
                    "type": "string",
                    "description": "Directory to save the MP3 file. Default: ~/Desktop",
                },
                "output_filename": {
                    "type": "string",
                    "description": "Custom filename. Auto-generated if not set.",
                },
                "speed": {
                    "type": "number",
                    "description": "Speech speed multiplier. Default: 1.0",
                },
            },
        },
    },
    {
        "name": "search_proximity",
        "description": "Find where one keyword appears near another keyword in audio. Useful for finding contextual discussions, e.g., 'startup' near 'funding'.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "audio_path": {
                    "type": "string",
                    "description": "Path to the audio file",
                },
                "keyword1": {
                    "type": "string",
                    "description": "Primary keyword to find",
                },
                "keyword2": {
                    "type": "string",
                    "description": "Secondary keyword that must appear nearby",
                },
                "max_distance": {
                    "type": "integer",
                    "description": "Maximum number of words between keywords. Default: 30",
                },
                "model_size": {
                    "type": "string",
                    "enum": [
                        "tiny",
                        "base",
                        "small",
                        "medium",
                        "large",
                    ],
                    "description": "Whisper model size. ALWAYS use tiny unless the user explicitly requests a different size. tiny is already highly accurate.",
                },
                "output": {
                    "type": "string",
                    "description": "Optional file path to save results. Use .csv for plain data or .xlsx for styled spreadsheets with bold headers and formatting.",
                },
            },
            "required": ["audio_path", "keyword1", "keyword2"],
        },
    },
    {
        "name": "identify_speakers",
        "description": "Identify who speaks when in audio.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "audio_path": {
                    "type": "string",
                    "description": "Path to the audio file",
                },
                "num_speakers": {
                    "type": "integer",
                    "description": "Number of speakers if known. Auto-detects if not set.",
                },
                "model_size": {
                    "type": "string",
                    "enum": [
                        "tiny",
                        "base",
                        "small",
                        "medium",
                        "large",
                    ],
                    "description": "Whisper model size. ALWAYS use tiny unless the user explicitly requests a different size. tiny is already highly accurate.",
                },
            },
            "required": ["audio_path"],
        },
    },
    {
        "name": "list_files",
        "description": "List media files in a directory.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "directory": {
                    "type": "string",
                    "description": "Directory path to search",
                },
                "pattern": {
                    "type": "string",
                    "description": "Glob pattern for matching files. Default: all common media formats",
                },
                "recursive": {
                    "type": "boolean",
                    "description": "Search subdirectories. Default: false",
                },
            },
            "required": ["directory"],
        },
    },
    {
        "name": "list_memories",
        "description": "List all stored transcriptions with their titles, durations, dates, and file paths to markdown files. Useful for browsing what has already been transcribed.",
        "inputSchema": {"type": "object", "properties": {}},
    },
    {
        "name": "memory_stats",
        "description": "View transcription memory statistics including number of stored files and total duration.",
        "inputSchema": {"type": "object", "properties": {}},
    },
    {
        "name": "clear_memory",
        "description": "Clear the transcription memory to free disk space.",
        "inputSchema": {"type": "object", "properties": {}},
    },
    {
        "name": "separate_audio",
        "description": "Separate audio into stems (vocals, drums, bass, other) using Meta's Demucs v4. Isolates vocals from music, background noise, and other sounds. Use this before transcription when audio has music, intros, or heavy background noise for dramatically cleaner results.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "audio_path": {
                    "type": "string",
                    "description": "Path to the audio file",
                },
                "vocals_only": {
                    "type": "boolean",
                    "description": "If true, only separate into vocals + no_vocals (faster). If false, separate into all 4 stems: vocals, drums, bass, other. Default: true",
                },
                "model": {
                    "type": "string",
                    "enum": ["htdemucs", "htdemucs_ft"],
                    "description": "Demucs model. htdemucs is the default (fast, great quality). htdemucs_ft is fine-tuned (slower, best quality). Default: htdemucs",
                },
            },
            "required": ["audio_path"],
        },
    },
    {
        "name": "search_memory",
        "description": "Search across ALL stored transcriptions. No audio_path needed, queries everything in memory. Default mode is 'keyword' (literal match). Use 'semantic' mode for meaning-based search.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "Search query. For keyword mode: a word or phrase to find literally. For semantic mode: a natural language description (e.g. 'discussion about funding challenges').",
                },
                "mode": {
                    "type": "string",
                    "enum": ["keyword", "semantic"],
                    "description": "Search mode. 'keyword' (default) finds segments containing the exact word/phrase. 'semantic' finds segments similar in meaning.",
                },
                "top_k": {
                    "type": "integer",
                    "description": "Number of results to return. Default: 10",
                },
                "output": {
                    "type": "string",
                    "description": "Optional file path to save results. Use .csv for plain data or .xlsx for styled spreadsheets with bold headers and formatting.",
                },
                "context_words": {
                    "type": "integer",
                    "description": "Words of context per result. Default: 25. Use 150 for full evidence blocks when Claude needs to answer a question. Semantic mode only.",
                },
                "dedup_seconds": {
                    "type": "number",
                    "description": "Merge matches within this many seconds of each other. Default: 0 (off). Semantic mode only.",
                },
            },
            "required": ["query"],
        },
    },
    {
        "name": "clip_export",
        "description": "Export a video clip from a URL for a specific time range. Downloads only the requested segment — not the full video. Perfect for extracting moments around keyword matches. Supports YouTube and 1000+ sites.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "url": {
                    "type": "string",
                    "description": "Video URL to extract clip from (YouTube, Vimeo, etc.)",
                },
                "start": {
                    "type": "number",
                    "description": "Start time in seconds",
                },
                "end": {
                    "type": "number",
                    "description": "End time in seconds",
                },
                "output_dir": {
                    "type": "string",
                    "description": "Directory to save the clip. Default: ~/Desktop",
                },
                "output_filename": {
                    "type": "string",
                    "description": "Custom filename for the clip (without extension). Auto-generated if not set.",
                },
            },
            "required": ["url", "start", "end"],
        },
    },
    {
        "name": "tag",
        "description": "Add, remove, or list tags on a transcription. Tags are broad topic categories (e.g. 'AI', 'Health', 'Music') that help organize and filter memories in the Web UI.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "cache_key": {
                    "type": "string",
                    "description": "The cache_key of the transcription to tag",
                },
                "action": {
                    "type": "string",
                    "description": "Action to perform: add, remove, or list",
                    "enum": ["add", "remove", "list"],
                },
                "tags": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Tag names to add or remove. Use 2-4 broad topic categories, not names of people or specific tools. E.g. ['AI', 'Startups'] not ['Greg Eisenberg', 'Claude Code']",
                },
            },
            "required": ["cache_key", "action"],
        },
    },
    {
        "name": "highlights",
        "description": "Export MP4 clips of specific moments. Two modes: auto (AI picks top moments by quotability and insight density) or focused (find moments matching a specific topic, person, or concept). Returns timestamps and text for each highlight, the calling agent decides which to export as clips.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "audio_path": {
                    "type": "string",
                    "description": "Path to the audio file (must be transcribed already)",
                },
                "query": {
                    "type": "string",
                    "description": "What to highlight. Omit for auto mode (top moments). Provide a topic, person, concept, or description for focused mode. Examples: 'product recommendations', 'heated debate moments', 'life advice'",
                },
                "top_k": {
                    "type": "number",
                    "description": "Number of highlights to return. Default: 5",
                },
                "model_size": {
                    "type": "string",
                    "description": "Whisper model size. ALWAYS use tiny unless the user explicitly requests a different size.",
                    "enum": [
                        "tiny",
                        "base",
                        "small",
                        "medium",
                        "large",
                    ],
                },
                "clip": {
                    "type": "boolean",
                    "description": "Export each highlight as an MP4 video clip. Requires the audio to have been downloaded from a URL. Default: false",
                },
                "clip_padding": {
                    "type": "number",
                    "description": "Seconds of padding around each highlight when exporting clips. Default: 15",
                },
                "context_words": {
                    "type": "number",
                    "description": "Words of context around each highlight. Default: 40",
                },
            },
            "required": ["audio_path"],
        },
    },
    {
        "name": "rebuild_graph",
        "description": "Rebuild Obsidian graph view data for all transcriptions. Migrates .md files to YAML frontmatter format, computes [[wikilinks]] between semantically related transcriptions, and generates MOC (Map of Content) hub files for tag clusters. Safe to run repeatedly. Run this once to upgrade existing memory for Obsidian.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "min_moc_members": {
                    "type": "number",
                    "description": "Minimum transcriptions per tag to generate a MOC file. Default: 3",
                },
            },
        },
    },
    {
        "name": "visual",
        "description": "Extract visual context from a video at moments that matter. Four modes: (1) Query mode (default): describe what you need visual context for and the tool finds those moments in the transcript, then extracts frames. (2) Auto mode: autonomously detects moments where the speaker implies visual content (UI actions, screen recordings, demonstrations). (3) Manual mode: extract frames at specific timestamps. (4) Assist mode: analyzes the transcript for visual gaps and returns time ranges where the user should provide their own screenshots (ideal for talking-head videos where the speaker describes a UI but doesn't show it). Frames are stored in augent memory alongside the transcription and embedded in the .md file as Obsidian wikilinks.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "video_path": {
                    "type": "string",
                    "description": "Path to a video file (MP4, MKV, etc). Can be output from clip_export.",
                },
                "url": {
                    "type": "string",
                    "description": "Video URL (YouTube, etc). Downloads the video automatically if video_path is not provided.",
                },
                "query": {
                    "type": "string",
                    "description": "What you need visual context for. The tool searches the transcript semantically and extracts frames at matching moments. Examples: 'connecting Gmail to the agent', 'the dashboard configuration', 'where he sets up the branching logic'.",
                },
                "timestamps": {
                    "type": "array",
                    "items": {"type": "number"},
                    "description": "Manual mode: list of timestamps (in seconds) to extract frames at. Overrides query and auto detection.",
                },
                "auto": {
                    "type": "boolean",
                    "description": "Auto mode: autonomously detect visual moments from transcript (UI actions, demonstrations, spatial references). Default: false. Ignored if query or timestamps are provided.",
                },
                "assist": {
                    "type": "boolean",
                    "description": "Assist mode: analyze the transcript for visual gaps and return time ranges where the user should provide their own screenshots. Ideal for talking-head videos or podcasts where the speaker describes a workflow or UI but the video doesn't show it. No frames are extracted — instead, returns structured gaps with time windows, transcript excerpts, and what kind of screenshot would help. Default: false.",
                },
                "model_size": {
                    "type": "string",
                    "description": "Whisper model size for transcription. Default: 'tiny'",
                    "enum": ["tiny", "base", "small", "medium", "large"],
                },
                "max_frames": {
                    "type": "number",
                    "description": "Maximum frames to extract. Default: 30",
                },
                "top_k": {
                    "type": "number",
                    "description": "Number of transcript matches to extract frames for in query mode. Default: 10",
                },
                "context_words": {
                    "type": "number",
                    "description": "Words of context around each match in query mode. Default: 40",
                },
                "clear": {
                    "type": "boolean",
                    "description": "Remove all previously extracted frames and the visual context .md for this video. Use to redo or clean up. Default: false",
                },
            },
            "required": [],
        },
    },
    {
        "name": "spaces",
        "description": "Download or live-record a Twitter/X Space. Three modes: (1) Pass url to start a download (returns recording_id), (2) Pass recording_id to check status, (3) Pass recording_id + stop=true to stop a live recording.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "url": {
                    "type": "string",
                    "description": "Twitter/X Space URL to download (e.g., https://x.com/i/spaces/1yNxaNvaMYQKj). Starts download in background.",
                },
                "output_dir": {
                    "type": "string",
                    "description": "Directory to save the audio file. Default: ~/Downloads",
                },
                "recording_id": {
                    "type": "string",
                    "description": "Check status of a previous download, or stop it when combined with stop=true.",
                },
                "stop": {
                    "type": "boolean",
                    "description": "Stop a live recording. Requires recording_id. Default: false",
                },
            },
            "required": [],
        },
    },
]

_active_recordings: dict = {}


def handle_tools_list(id: Any) -> None:
    """Handle tools/list request — filters out disabled tools from config."""
    from .config import get_config

    disabled = set(get_config().get("disabled_tools", []))
    tools = [t for t in _ALL_TOOLS if t["name"] not in disabled]
    send_response(
        {
            "jsonrpc": "2.0",
            "id": id,
            "result": {"tools": tools},
        }
    )


def handle_tools_call(id: Any, params: dict) -> None:
    """Handle tools/call request."""
    from .config import get_config

    tool_name = params.get("name")
    arguments = params.get("arguments", {})

    disabled = set(get_config().get("disabled_tools", []))
    if tool_name in disabled:
        send_error(id, -32602, f"Tool is disabled: {tool_name}")
        return

    try:
        if tool_name == "download_audio":
            result = handle_download_audio(arguments)
        elif tool_name == "transcribe_audio":
            result = handle_transcribe_audio(arguments)
        elif tool_name == "search_audio":
            result = handle_search_audio(arguments)
        elif tool_name == "deep_search":
            result = handle_deep_search(arguments)
        elif tool_name == "take_notes":
            result = handle_take_notes(arguments)
        elif tool_name == "chapters":
            result = handle_chapters(arguments)
        elif tool_name == "batch_search":
            result = handle_batch_search(arguments)
        elif tool_name == "text_to_speech":
            result = handle_text_to_speech(arguments)
        elif tool_name == "search_proximity":
            result = handle_search_proximity(arguments)
        elif tool_name == "identify_speakers":
            result = handle_identify_speakers(arguments)
        elif tool_name == "list_files":
            result = handle_list_files(arguments)
        elif tool_name == "list_memories":
            result = handle_list_memories(arguments)
        elif tool_name == "memory_stats":
            result = handle_memory_stats(arguments)
        elif tool_name == "clear_memory":
            result = handle_clear_memory(arguments)
        elif tool_name == "search_memory":
            result = handle_search_memory(arguments)
        elif tool_name == "separate_audio":
            result = handle_separate_audio(arguments)
        elif tool_name == "clip_export":
            result = handle_clip_export(arguments)
        elif tool_name == "highlights":
            result = handle_highlights(arguments)
        elif tool_name == "tag":
            result = handle_tag(arguments)
        elif tool_name == "rebuild_graph":
            result = handle_rebuild_graph(arguments)
        elif tool_name == "visual":
            result = handle_visual(arguments)
        elif tool_name == "spaces":
            result = handle_spaces(arguments)
        else:
            send_error(id, -32602, f"Unknown tool: {tool_name}")
            return

        send_response(
            {
                "jsonrpc": "2.0",
                "id": id,
                "result": {
                    "content": [{"type": "text", "text": json.dumps(result, indent=2)}]
                },
            }
        )

    except FileNotFoundError as e:
        send_error(id, -32602, str(e))
    except ValueError as e:
        send_error(id, -32602, str(e))
    except Exception as e:
        send_error(id, -32603, f"Error: {str(e)}")


def handle_download_audio(arguments: dict) -> dict:
    """Handle download_audio tool call."""
    import os
    import re
    import shutil
    import subprocess

    from .config import get_config

    cfg = get_config()
    url = arguments.get("url")
    output_dir = arguments.get("output_dir", os.path.expanduser(cfg["output_dir"]))

    if not url:
        raise ValueError("Missing required parameter: url")

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Find yt-dlp — prefer brew (stays current) over pip
    ytdlp = shutil.which(
        "yt-dlp", path="/opt/homebrew/bin:/usr/local/bin"
    ) or shutil.which("yt-dlp")
    if not ytdlp:
        raise RuntimeError("yt-dlp not found. Install with: brew install yt-dlp")

    # Check for aria2c (optional but recommended)
    has_aria2c = shutil.which("aria2c") is not None

    # Build command — bestaudio/best fallback handles YouTube SABR streaming
    cmd = [
        ytdlp,
        "-f",
        "bestaudio/best",
        "--concurrent-fragments",
        "4",
        "--no-playlist",
        "--restrict-filenames",
        "-o",
        f"{output_dir}/%(title)s [%(id)s].%(ext)s",
        "--print",
        "after_move:filepath",  # Print the final file path
    ]

    if has_aria2c:
        cmd.extend(
            ["--downloader", "aria2c", "--downloader-args", "aria2c:-x 16 -s 16 -k 1M"]
        )

    cmd.append(url)

    # Run download
    result = subprocess.run(cmd, capture_output=True, text=True)

    if result.returncode != 0:
        error_msg = result.stderr.strip() or "Download failed"
        raise RuntimeError(f"Download failed: {error_msg}")

    # Extract the output file path from stdout
    output_lines = result.stdout.strip().split("\n")
    output_file = output_lines[-1] if output_lines else None

    # Get file info if available
    file_info = {}
    if output_file and os.path.exists(output_file):
        file_size = os.path.getsize(output_file)
        file_info = {
            "path": output_file,
            "filename": os.path.basename(output_file),
            "size_mb": round(file_size / (1024 * 1024), 2),
        }

    # Register source URL so transcription can attach it to memory
    if output_file and os.path.exists(output_file):
        _downloaded_urls[os.path.abspath(output_file)] = url
        # Persist source URL permanently (survives restarts)
        try:
            from .memory import get_transcription_memory

            get_transcription_memory().save_source_url(
                os.path.abspath(output_file), url
            )
        except Exception:
            pass

    return {
        "success": True,
        "url": url,
        "output_dir": output_dir,
        "file": file_info,
        "aria2c_used": has_aria2c,
        "message": (
            f"Audio downloaded to {output_file}" if output_file else "Download complete"
        ),
    }


def handle_search_audio(arguments: dict) -> dict:
    """Handle search_audio tool call."""
    from .config import get_config

    cfg = get_config()
    audio_path = arguments.get("audio_path")
    keywords = arguments.get("keywords", [])
    model_size = arguments.get("model_size", cfg["model_size"])
    include_full = arguments.get("include_full_text", False)
    output = arguments.get("output")
    clip = arguments.get("clip", False)
    clip_padding = arguments.get("clip_padding", cfg["clip_padding"])

    if not audio_path:
        raise ValueError("Missing required parameter: audio_path")
    if not keywords:
        raise ValueError("Missing required parameter: keywords")

    if include_full:
        result = search_audio_full(audio_path, keywords, model_size=model_size)
    else:
        result = search_audio(audio_path, keywords, model_size=model_size)

    result["model_used"] = model_size

    # Look up source URL for YouTube timestamp linking
    source_url = _downloaded_urls.get(os.path.abspath(audio_path), "")
    if not source_url:
        from .memory import get_transcription_memory

        mem = get_transcription_memory()
        source_url = mem.get_source_url(audio_path, model_size)
        if not source_url:
            source_url = mem.get_source_url_by_hash(audio_path)

    # Add YouTube links to keyword matches
    if source_url and _extract_youtube_id(source_url):
        for _kw, matches in result.items():
            if isinstance(matches, list):
                for m in matches:
                    secs = m.get("timestamp_seconds", 0)
                    if secs:
                        m["youtube_link"] = _youtube_timestamp_link(source_url, secs)

    # Write output file if requested
    if output:
        # Flatten grouped results into rows
        rows = []
        for kw, matches in result.items():
            if isinstance(matches, list):
                for m in matches:
                    row = {
                        "keyword": kw,
                        "timestamp": m.get("timestamp", ""),
                        "timestamp_seconds": m.get("timestamp_seconds", 0),
                        "snippet": m.get("snippet", ""),
                    }
                    if m.get("youtube_link"):
                        row["youtube_link"] = m["youtube_link"]
                    rows.append(row)
        if rows:
            cols = ["keyword", "timestamp", "snippet"]
            if any(r.get("youtube_link") for r in rows):
                cols.append("youtube_link")
            result["output_path"] = _write_output_file(
                output,
                rows,
                columns=cols,
                bold_columns=["keyword", "timestamp"],
            )

    # Export clips around matches if requested
    if clip and source_url:
        timestamps = []
        for _kw, matches in result.items():
            if isinstance(matches, list):
                for m in matches:
                    ts = m.get("timestamp_seconds", 0)
                    if ts:
                        timestamps.append(float(ts))
        if timestamps:
            result["clips"] = _export_clips_for_matches(
                source_url, timestamps, padding=clip_padding
            )
        else:
            result["clips"] = []
            result["clip_note"] = "No matches with timestamps to clip."
    elif clip and not source_url:
        result["clips"] = []
        result["clip_note"] = (
            "No source URL found for this audio file. "
            "Clips can only be exported when the audio was downloaded from a URL."
        )

    return result


def handle_transcribe_audio(arguments: dict) -> dict:
    """Handle transcribe_audio tool call."""
    import subprocess
    import tempfile

    from .config import get_config

    cfg = get_config()
    audio_path = arguments.get("audio_path")
    model_size = arguments.get("model_size", cfg["model_size"])
    start = arguments.get("start")
    duration = arguments.get("duration")
    output = arguments.get("output")
    translated_text = arguments.get("translated_text")

    if not audio_path:
        raise ValueError("Missing required parameter: audio_path")

    # If only storing a translation, do that and return early (no re-transcription)
    if translated_text:
        from .memory import get_transcription_memory

        memory = get_transcription_memory()
        md_path = memory.store_translation(audio_path, model_size, translated_text)
        if md_path:
            return {
                "status": "translation_stored",
                "translated_md_path": md_path,
                "message": "English translation saved alongside original transcription.",
            }
        else:
            raise ValueError(
                "No existing transcription found for this audio_path + model_size. "
                "Transcribe the audio first, then store the translation."
            )

    # If start or duration specified, trim audio with ffmpeg first
    trimmed_path = None
    if start is not None or duration is not None:
        tmp = tempfile.NamedTemporaryFile(suffix=".webm", delete=False)
        trimmed_path = tmp.name
        tmp.close()
        cmd = ["ffmpeg", "-y", "-i", audio_path]
        if start is not None:
            cmd.extend(["-ss", str(start)])
        if duration is not None:
            cmd.extend(["-t", str(duration)])
        cmd.extend(["-vn", "-acodec", "copy", trimmed_path])
        subprocess.run(cmd, capture_output=True, check=True)
        audio_path = trimmed_path

    # Resolve the original audio path (before trimming) for URL lookup
    original_audio_path = arguments.get("audio_path")

    try:
        result = transcribe_audio(audio_path, model_size)
    finally:
        # Clean up temp file
        if trimmed_path and os.path.exists(trimmed_path):
            os.remove(trimmed_path)

    # Attach source URL to memory if this file was downloaded via download_audio
    source_url = _downloaded_urls.get(os.path.abspath(original_audio_path), "")
    if not source_url:
        # Check persisted source URLs (survives restarts)
        from .memory import get_transcription_memory

        source_url = get_transcription_memory().get_source_url_by_hash(
            original_audio_path
        )
    if source_url:
        from .memory import get_transcription_memory

        get_transcription_memory().update_source_url(
            original_audio_path, model_size, source_url
        )

    # Build per-segment timestamps
    segments = []
    for seg in result.get("segments", []):
        s = seg["start"]
        # Offset timestamps back if start was specified
        if start:
            s += start
        e = seg["end"]
        if start:
            e += start
        minutes_s, secs_s = int(s // 60), int(s % 60)
        seg_dict = {
            "start": round(s, 1),
            "end": round(e, 1),
            "timestamp": f"{minutes_s}:{secs_s:02d}",
            "text": seg["text"].strip(),
        }
        yt_link = _youtube_timestamp_link(source_url, s)
        if yt_link:
            seg_dict["youtube_link"] = yt_link
        segments.append(seg_dict)

    # Cap response size to prevent token overflow in Claude Code.
    # For large transcriptions, truncate text and suggest using output param.
    full_text = result["text"]
    max_chars = 12000
    max_segments = 200
    truncated = len(full_text) > max_chars or len(segments) > max_segments

    if truncated:
        capped_text = (
            full_text[:max_chars]
            + "\n\n[... truncated — use the output parameter to export full transcription to .csv or .xlsx ...]"
        )
        capped_segments = segments[:max_segments]
    else:
        capped_text = full_text
        capped_segments = segments

    response = {
        "text": capped_text,
        "language": result["language"],
        "duration": result["duration"],
        "duration_formatted": f"{int(result['duration'] // 60)}:{int(result['duration'] % 60):02d}",
        "segments": capped_segments,
        "segment_count": len(segments),
        "cached": result.get("cached", False),
        "model_used": model_size,
    }

    if result["language"] != "en":
        lang = result["language"]
        response["translation_available"] = True
        response["MANDATORY_TRANSLATION_OFFER"] = (
            f"After completing your response, you MUST append exactly:\n\n"
            f"---\n"
            f"This audio is in **{lang}**. Would you like me to translate it to English and store it in your augent memory?\n\n"
            f"If the user says yes, translate the transcription yourself and call "
            f'transcribe_audio(audio_path="{original_audio_path}", translated_text="<your full english translation>") '
            f"to store it."
        )

    if truncated:
        response["truncated"] = True
        response["full_segment_count"] = len(segments)
        response["hint"] = (
            "Response was truncated to prevent overflow. Use the output parameter (e.g. output: '~/Desktop/transcript.csv') to get the full transcription."
        )

    # Write output file if requested
    if output:
        cols = ["timestamp", "text"]
        if any(s.get("youtube_link") for s in segments):
            cols.append("youtube_link")
        response["output_path"] = _write_output_file(
            output,
            segments,
            columns=cols,
            bold_columns=["timestamp"],
        )

    # Semantic tagging — assign existing tags based on content similarity
    try:
        from .embeddings import semantic_tag
        from .memory import get_transcription_memory

        _mem = get_transcription_memory()
        _audio_hash = _mem.hash_audio_file(audio_path)
        _ck = f"{_audio_hash}:{model_size}"
        response["cache_key"] = _ck

        # Run semantic tagging first (matches against existing tags)
        transcript_text = response.get("text", "")
        semantic_tag(_ck, transcript_text)

        # Then hint Claude to add new categories if needed
        existing_tags = _mem.get_tags(_ck)
        all_tags = _mem.get_all_tags_with_counts()
        existing_names = [t["name"] for t in all_tags]
        assigned_names = [t["name"] for t in existing_tags]
        hint = (
            f"Silently call the `tag` tool with "
            f'cache_key: "{_ck}", action: "add", and 2-4 broad topic tags. '
            f"Use short, high-level categories — not people's names, not specific tools, not subtopics. "
        )
        if assigned_names:
            hint += f"Already assigned by semantic matching: {assigned_names}. Only add tags if a major topic is missing. "
        if existing_names:
            hint += f"Existing tags in this library: {existing_names}. Reuse these where they fit before creating new ones. "
        hint += "Do not mention tagging to the user."
        response["tagging_hint"] = hint
    except Exception:
        pass

    return response


def handle_search_proximity(arguments: dict) -> dict:
    """Handle search_proximity tool call."""
    from .config import get_config

    cfg = get_config()
    audio_path = arguments.get("audio_path")
    keyword1 = arguments.get("keyword1")
    keyword2 = arguments.get("keyword2")
    max_distance = arguments.get("max_distance", 30)
    model_size = arguments.get("model_size", cfg["model_size"])
    output = arguments.get("output")

    if not audio_path:
        raise ValueError("Missing required parameter: audio_path")
    if not keyword1:
        raise ValueError("Missing required parameter: keyword1")
    if not keyword2:
        raise ValueError("Missing required parameter: keyword2")

    matches = search_audio_proximity(
        audio_path, keyword1, keyword2, max_distance=max_distance, model_size=model_size
    )

    # Add YouTube links if source is YouTube
    source_url = _downloaded_urls.get(os.path.abspath(audio_path), "")
    if not source_url:
        from .memory import get_transcription_memory

        mem = get_transcription_memory()
        source_url = mem.get_source_url(audio_path, model_size)
        if not source_url:
            source_url = mem.get_source_url_by_hash(audio_path)

    if source_url and _extract_youtube_id(source_url):
        for m in matches:
            secs = m.get("timestamp_seconds", 0)
            yt_link = _youtube_timestamp_link(source_url, secs)
            if yt_link:
                m["youtube_link"] = yt_link

    result = {
        "query": f"'{keyword1}' within {max_distance} words of '{keyword2}'",
        "match_count": len(matches),
        "matches": matches,
        "model_used": model_size,
    }

    # Write output file if requested
    if output and matches:
        cols = ["timestamp", "snippet"]
        if any(m.get("youtube_link") for m in matches):
            cols.append("youtube_link")
        result["output_path"] = _write_output_file(
            output,
            matches,
            columns=cols,
            bold_columns=["timestamp"],
        )

    return result


def handle_batch_search(arguments: dict) -> dict:
    """Handle batch_search tool call."""
    import os
    from concurrent.futures import ThreadPoolExecutor, as_completed

    from .config import get_config

    cfg = get_config()
    audio_paths = arguments.get("audio_paths", [])
    keywords = arguments.get("keywords", [])
    model_size = arguments.get("model_size", cfg["model_size"])
    workers = arguments.get("workers", 2)

    if not audio_paths:
        raise ValueError("Missing required parameter: audio_paths")
    if not keywords:
        raise ValueError("Missing required parameter: keywords")

    # Validate all paths exist
    valid_paths = []
    errors = []
    for path in audio_paths:
        if os.path.exists(path):
            valid_paths.append(path)
        else:
            errors.append({"path": path, "error": "File not found"})

    results = {}

    def process_file(path):
        try:
            return path, search_audio(path, keywords, model_size=model_size)
        except Exception as e:
            return path, {"error": str(e)}

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(process_file, p): p for p in valid_paths}
        for future in as_completed(futures):
            path, result = future.result()
            results[path] = result

    # Aggregate stats
    total_matches = 0
    for _path, file_results in results.items():
        if "error" not in file_results:
            for _kw, matches in file_results.items():
                total_matches += len(matches)

    return {
        "files_processed": len(valid_paths),
        "files_with_errors": len(errors),
        "total_matches": total_matches,
        "results": results,
        "errors": errors if errors else None,
        "model_used": model_size,
    }


def handle_list_files(arguments: dict) -> dict:
    """Handle list_files tool call."""
    import glob as glob_module
    import os

    DEFAULT_PATTERNS = [
        "*.mp3",
        "*.m4a",
        "*.wav",
        "*.webm",
        "*.mp4",
        "*.mkv",
        "*.ogg",
        "*.flac",
    ]

    directory = arguments.get("directory")
    pattern = arguments.get("pattern")
    recursive = arguments.get("recursive", False)

    if not directory:
        raise ValueError("Missing required parameter: directory")

    if not os.path.isdir(directory):
        raise ValueError(f"Directory not found: {directory}")

    # Build search pattern(s)
    files = []
    if pattern:
        patterns = [pattern]
    else:
        patterns = DEFAULT_PATTERNS

    for p in patterns:
        if recursive:
            search_pattern = os.path.join(directory, "**", p)
            files.extend(glob_module.glob(search_pattern, recursive=True))
        else:
            search_pattern = os.path.join(directory, p)
            files.extend(glob_module.glob(search_pattern))

    # Deduplicate and sort
    files = sorted(set(files))

    # Get file info
    audio_files = []
    for f in files:
        try:
            size = os.path.getsize(f)
            audio_files.append(
                {
                    "path": f,
                    "name": os.path.basename(f),
                    "size_mb": round(size / (1024 * 1024), 2),
                }
            )
        except OSError:
            pass

    return {
        "directory": directory,
        "pattern": pattern,
        "recursive": recursive,
        "count": len(audio_files),
        "files": audio_files,
    }


def handle_memory_stats(arguments: dict) -> dict:
    """Handle memory_stats tool call."""
    return get_memory_stats()


def handle_clear_memory(arguments: dict) -> dict:
    """Handle clear_memory tool call."""
    count = clear_memory()
    return {"cleared": count, "message": f"Cleared {count} stored transcription(s)"}


def handle_list_memories(arguments: dict) -> dict:
    """Handle list_memories tool call."""
    entries = list_memories()
    return {
        "count": len(entries),
        "transcriptions": entries,
        "message": f"Found {len(entries)} stored transcription(s)",
    }


def _get_style_instruction(
    style: str,
    read_aloud: bool = False,
    output_dir: str = "~/Desktop",
    safe_title: str = "",
    txt_path: str = "",
) -> str:
    """Return formatting instructions for a given note style."""

    base_prefix = (
        "IMPORTANT: You MUST now format the notes and save them by calling take_notes again with save_content. "
        "Do NOT use the Write or Edit tools on notes files — ALWAYS use take_notes(save_content=...) for both initial save and any subsequent edits. "
        "Do NOT leave the raw transcription as-is. Do NOT include YAML frontmatter — it is added automatically. "
        "For any follow-up tool calls (chapters, search, deep_search, etc.), use the audio_path field from this response — do NOT guess the filename. "
    )
    base_suffix = (
        '\n\nSave the final notes by calling: take_notes(save_content="<your formatted notes>"). '
        "Do NOT use the Write tool."
    )

    if read_aloud:
        import os as _os
        import shutil as _shutil

        _obsidian_installed = _os.path.exists("/Applications/Obsidian.app") or bool(
            _shutil.which("obsidian")
        )
        audio_filename = f"{safe_title}.mp3" if safe_title else "notes_audio.mp3"
        if _obsidian_installed:
            embed_instruction = (
                f"After TTS completes, prepend ![[{audio_filename}]] on the very first line "
                f"(before the title) and `> Press Cmd+E before playing — prevents audio from pausing on scroll` "
                "on the line after the embed, then save by calling take_notes(save_content=...) with the full updated content. "
            )
        else:
            embed_instruction = ""
        base_suffix = (
            '\n\nSave the final notes by calling: take_notes(save_content="<your formatted notes>"). '
            "Do NOT use the Write tool. "
            "THEN: Take the notes you just wrote — SKIP the title, source URL, duration, date, and any metadata lines at the top. Start from the first real content section heading. Take that content and "
            "strip the markdown formatting (remove #, **, -, >, ![], ---, callout syntax, links) "
            "so it reads as plain text. Keep every word and all the information exactly as written — "
            "do NOT rewrite or summarize, just clean the formatting so TTS can read it naturally. "
            "Section headers become spoken section titles. "
            "Run the text_to_speech tool with that spoken script, "
            f'output_dir="{output_dir}", output_filename="{audio_filename}". '
            + embed_instruction
        )

    styles = {
        "tldr": (
            "Create the shortest possible summary. Must fit on one screen.\n"
            "\n"
            "FORMAT:\n"
            "- Title as a top header\n"
            "- Source URL | Duration | Date on one line\n"
            "- ---\n"
            "- One 2-3 sentence overview paragraph\n"
            "- 5-8 bullet points max, each one line\n"
            "- **Bold** the single most important term or name in each bullet\n"
            "- No sections, no headers, no callouts, no quotes — just clean bullets\n"
            "- End with one bold takeaway line\n"
        ),
        "notes": (
            "Create clean, structured notes with clear hierarchy.\n"
            "\n"
            "FORMAT:\n"
            "- Title as a top header\n"
            "- Metadata block: Source URL, Duration, Date\n"
            "- ---\n"
            "- 3-6 section headers based on the main topics\n"
            "- Nested bullet points under each section (2 levels max)\n"
            "- **Bold** key terms and names throughout\n"
            "- Short paragraphs only — never more than 3 lines\n"
            "- One > blockquote if there's a standout quote worth preserving\n"
            "- Keep it scannable — someone should grasp the content in 60 seconds\n"
        ),
        "highlight": (
            "Create formatted notes with visual emphasis on key insights.\n"
            "\n"
            "FORMAT:\n"
            "- Title as a top header\n"
            "- Metadata block: Source URL, Duration, Date\n"
            "- ---\n"
            "- Section headers for each major topic\n"
            "- Nested bullet points with **bold key terms**\n"
            "- Use > [!tip] callout blocks for the 2-4 most important insights\n"
            "- Use > [!info] callout blocks for definitions or context\n"
            "- Use > blockquotes with timestamps for 2-3 key direct quotes\n"
            "- Use **bold** and *italic* generously for emphasis\n"
            "- Add a --- separator between major sections\n"
            "- End with a Key Takeaways section using a > [!summary] callout\n"
        ),
        "eye-candy": (
            "Create the most visually rich, beautifully formatted notes possible. "
            "Every section should be a visual experience — the reader should absorb "
            "the content by scanning, not reading.\n"
            "\n"
            "FORMAT:\n"
            "- Title as a top header\n"
            "- Metadata block: Source URL, Duration, Date\n"
            "- ---\n"
            "- Section headers for every topic shift\n"
            "- Nested bullet points (up to 3 levels) with **bold** and *italic*\n"
            "- > [!tip] callout blocks for key insights (use liberally, 4-6 throughout)\n"
            "- > [!info] callout blocks for context, background, definitions\n"
            "- > [!warning] callout blocks for common mistakes or misconceptions\n"
            "- > [!example] callout blocks for concrete examples mentioned\n"
            "- > blockquotes with timestamps for 3-5 standout direct quotes\n"
            "- Tables anywhere a comparison or list of items is discussed\n"
            "- --- separators between major sections\n"
            "- Checklists (- [ ]) for any action items or recommendations\n"
            "- End with:\n"
            "  1. A > [!summary] Key Takeaways callout with numbered list\n"
            "  2. A table of Related Topics / Further Reading if applicable\n"
            "\n"
            "The goal: someone opens this file in Obsidian and says 'wow'.\n"
        ),
        "quiz": (
            "Generate a multiple-choice quiz from the content. Do NOT write notes.\n"
            "\n"
            "FORMAT:\n"
            "- Title as a top header with 'Quiz' appended\n"
            "- Metadata block: Source URL, Duration, Date\n"
            "- ---\n"
            "- 10-15 multiple-choice questions\n"
            "- Each question MUST follow this EXACT structure:\n"
            "\n"
            "### 1. **Question text here?**\n"
            "\n"
            "A) First option\n"
            "B) Second option\n"
            "C) Third option\n"
            "D) Fourth option\n"
            "\n"
            "- ---\n"
            "- Answer Key section at the bottom with this EXACT format:\n"
            "\n"
            "## Answer Key\n"
            "\n"
            "**1. B** — Explanation of why B is correct.\n"
            "**2. A** — Explanation of why A is correct.\n"
            "**3. D** — Explanation of why D is correct.\n"
            "\n"
            "Each answer on its own line. Bold number and letter, em dash, then explanation. No grouping, no compact rows, no bullet lists.\n"
            "\n"
            "Questions should test real understanding, not trivial details.\n"
        ),
    }

    body = styles.get(style, styles["notes"])
    return base_prefix + body + base_suffix


_last_notes_path = None
_last_notes_metadata = (
    {}
)  # Stored during initial take_notes for frontmatter in save_content


def handle_take_notes(arguments: dict) -> dict:
    """Handle take_notes tool call - download, transcribe, save .md to Desktop."""
    import os
    import re

    global _last_notes_path, _last_notes_metadata

    # --- Save mode: write formatted notes to a file ---
    save_content = arguments.get("save_content")
    if save_content is not None:
        output_path = arguments.get("output_path")
        if output_path:
            output_path = os.path.expanduser(output_path)
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            _last_notes_path = output_path
        if not _last_notes_path:
            raise ValueError(
                "No previous take_notes path. Call take_notes with a url first, or provide output_path."
            )
        # Post-process: convert plain A)/B)/C)/D) answer lines to checkbox syntax
        had_bare = bool(re.search(r"^\s*[A-D]\)", save_content, flags=re.MULTILINE))
        save_content = re.sub(
            r"^(\s*)([A-D]\))", r"- [ ] \2", save_content, flags=re.MULTILINE
        )
        has_checkboxes = "- [ ]" in save_content

        # Strip any existing frontmatter from save_content (we generate our own)
        if save_content.lstrip().startswith("---\n"):
            stripped = save_content.lstrip()
            end_idx = stripped.find("\n---\n", 4)
            if end_idx != -1:
                save_content = stripped[end_idx + 5 :]

        # Auto-prepend YAML frontmatter for Obsidian graph view
        if _last_notes_metadata:
            from .memory import TranscriptionMemory

            meta = _last_notes_metadata
            # Build wikilink to source transcription
            extra = {}
            if meta.get("source_transcription_filename"):
                extra["source_transcription"] = (
                    f'"[[{meta["source_transcription_filename"]}]]"'
                )
            if meta.get("style"):
                extra["style"] = meta["style"]

            frontmatter = TranscriptionMemory._build_frontmatter(
                title=meta.get("title", ""),
                tags=meta.get("tags", []),
                source_url=meta.get("source_url", ""),
                duration=meta.get("duration_formatted", ""),
                language=meta.get("language", ""),
                date=meta.get("date", ""),
                type_="notes",
                extra=extra,
            )
            save_content = frontmatter + "\n" + save_content

        with open(_last_notes_path, "w", encoding="utf-8") as f:
            f.write(save_content)
        return {
            "success": True,
            "saved_to": _last_notes_path,
            "size": len(save_content),
            "debug_checkbox": {
                "had_bare_options": had_bare,
                "has_checkboxes_after": has_checkboxes,
            },
        }

    from .config import get_config

    cfg = get_config()
    url = arguments.get("url")
    output_dir = arguments.get(
        "output_dir", os.path.expanduser(cfg["notes_output_dir"])
    )
    model_size = arguments.get("model_size", cfg["model_size"])
    style = arguments.get("style", "notes")
    read_aloud = arguments.get("read_aloud", False)
    visual_query = arguments.get("visual")

    if not url:
        raise ValueError("Missing required parameter: url")

    os.makedirs(output_dir, exist_ok=True)

    # Step 1: Download audio to ~/Downloads
    download_result = handle_download_audio({"url": url})

    if not download_result.get("success"):
        raise RuntimeError(
            "Download failed: " + download_result.get("message", "unknown error")
        )

    file_info = download_result.get("file", {})
    if not file_info.get("path"):
        raise RuntimeError("Download succeeded but output file not found")
    audio_path = file_info["path"]
    title = os.path.splitext(file_info["filename"])[0]

    # Step 2: Transcribe
    result = transcribe_audio(audio_path, model_size)
    text = result["text"]
    duration = result["duration"]

    # Attach source URL to memory
    from .memory import get_transcription_memory

    get_transcription_memory().update_source_url(audio_path, model_size, url)

    # Step 3: Save raw transcription as .md on Desktop
    # Clean title for filename (remove special chars)
    safe_title = re.sub(r"[^\w\s\-]", "", title)
    safe_title = re.sub(r"\s+", " ", safe_title).strip()
    if not safe_title:
        safe_title = "notes"
    md_filename = f"{safe_title}.md"
    md_path = os.path.join(output_dir, md_filename)

    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"Source: {url}\n")
        f.write(f"Duration: {int(duration // 60)}:{int(duration % 60):02d}\n")
        f.write(f"Title: {title}\n")
        f.write("=" * 60 + "\n\n")
        f.write(text)

    _last_notes_path = md_path

    # Style-specific formatting instructions
    instruction = _get_style_instruction(
        style,
        read_aloud=read_aloud,
        output_dir=output_dir,
        safe_title=safe_title,
        txt_path=md_path,
    )

    # Cap transcription in response to prevent token overflow in Claude Code.
    # Full transcript is already saved to md_path — Claude reads the file.
    max_chars = 12000
    if len(text) > max_chars:
        truncated_text = (
            text[:max_chars]
            + f"\n\n[... truncated — full transcript saved to {md_path} — read the file for complete text ...]"
        )
    else:
        truncated_text = text

    response = {
        "success": True,
        "write_to": md_path,
        "audio_path": audio_path,
        "instruction": instruction,
        "title": title,
        "duration": duration,
        "duration_formatted": f"{int(duration // 60)}:{int(duration % 60):02d}",
        "language": result["language"],
        "cached": result.get("cached", False),
        "model_used": model_size,
        "style": style,
        "transcription": truncated_text,
    }

    label = "Quiz" if style == "quiz" else "Notes"
    if result["language"] != "en":
        lang = result["language"]
        response["translation_available"] = True
        response["MANDATORY_TRANSLATION_OFFER"] = (
            f"After saving the notes, you MUST append exactly:\n\n"
            f"---\n"
            f"This audio is in **{lang}**. Would you like me to translate it to English "
            f"and store it in your augent memory?\n\n"
            f"If the user says yes, translate the transcription yourself and call "
            f'transcribe_audio(audio_path="{audio_path}", translated_text="<your full english translation>") '
            f"to store it."
        )
    else:
        response[
            "instruction"
        ] += f"\n\nAfter saving, respond ONLY with: Done. {label} saved to ~/Desktop/<filename>"

    # Visual context hint — let Claude know visual is available for this video
    response["visual_hint"] = (
        f"If the user needs visual context (screenshots of UI, dashboards, workflows shown in the video), "
        f'use the `visual` tool with url: "{url}" and a query describing what they need to see. '
        f'Example: visual(url="{url}", query="setting up the automation")'
    )

    # Semantic tagging — assign existing tags based on content similarity
    try:
        from .embeddings import semantic_tag
        from .memory import get_transcription_memory

        _mem = get_transcription_memory()
        _audio_hash = _mem.hash_audio_file(audio_path)
        _ck = f"{_audio_hash}:{model_size}"
        response["cache_key"] = _ck

        # Run semantic tagging first (matches against existing tags)
        semantic_tag(_ck, text)

        # Then hint Claude to add new categories if needed
        existing_tags = _mem.get_tags(_ck)
        all_tags = _mem.get_all_tags_with_counts()
        existing_names = [t["name"] for t in all_tags]
        assigned_names = [t["name"] for t in existing_tags]
        hint = (
            f"Silently call the `tag` tool with "
            f'cache_key: "{_ck}", action: "add", and 2-4 broad topic tags. '
            f"Use short, high-level categories — not people's names, not specific tools, not subtopics. "
        )
        if assigned_names:
            hint += f"Already assigned by semantic matching: {assigned_names}. Only add tags if a major topic is missing. "
        if existing_names:
            hint += f"Existing tags in this library: {existing_names}. Reuse these where they fit before creating new ones. "
        hint += "Do not mention tagging to the user."
        response["tagging_hint"] = hint

        # Store metadata for frontmatter generation in save_content
        _source_sanitized = _mem._sanitize_filename(_mem._title_from_path(audio_path))
        from datetime import datetime as _dt

        _last_notes_metadata = {
            "title": title,
            "tags": assigned_names,
            "source_url": url,
            "duration_formatted": f"{int(duration // 60)}:{int(duration % 60):02d}",
            "language": result["language"],
            "date": _dt.now().strftime("%Y-%m-%d"),
            "style": style,
            "source_transcription_filename": _source_sanitized,
            "cache_key": _ck,
        }
    except Exception:
        pass

    # Ensure metadata is set even if tagging failed
    if not _last_notes_metadata:
        from datetime import datetime as _dt2

        _last_notes_metadata = {
            "title": title,
            "tags": [],
            "source_url": url,
            "duration_formatted": f"{int(duration // 60)}:{int(duration % 60):02d}",
            "language": result["language"],
            "date": _dt2.now().strftime("%Y-%m-%d"),
            "style": style,
            "source_transcription_filename": "",
            "cache_key": "",
        }

    # Visual context: download video and extract frames in one shot
    if visual_query:
        try:
            visual_result = handle_visual(
                {
                    "url": url,
                    "query": visual_query,
                    "model_size": model_size,
                }
            )
            frames_data = [
                {
                    "filename": f["filename"],
                    "timestamp": f["timestamp_formatted"],
                    "transcript": f.get("transcript", "")[:150],
                }
                for f in visual_result["frames"]
            ]
            response["visual"] = {
                "frame_count": visual_result["frame_count"],
                "frames_dir": visual_result["frames_dir"],
                "frames": frames_data,
            }
            # Tell Claude to embed frames inline in the notes
            if frames_data:
                embed_lines = []
                for fd in frames_data:
                    embed_lines.append(
                        f"- At {fd['timestamp']}: ![[{fd['filename']}]] (context: {fd['transcript'][:80]})"
                    )
                response["visual_embed_instruction"] = (
                    "IMPORTANT: When formatting the notes, embed these visual frames inline "
                    "at the relevant sections using Obsidian wikilink syntax. Place each "
                    "![[filename.png]] embed on its own line right after the section it relates to. "
                    "Do NOT create a separate visual section — weave them naturally into the notes.\n\n"
                    "Frames to embed:\n" + "\n".join(embed_lines)
                )
        except Exception as e:
            response["visual_error"] = str(e)

    return response


def handle_identify_speakers(arguments: dict) -> dict:
    """Handle identify_speakers tool call."""
    try:
        from .speakers import identify_speakers
    except ImportError as err:
        raise RuntimeError(
            "Missing dependencies: pyannote-audio. "
            "Install with: pip install augent[speakers]\n"
            "Then run: curl -fsSL https://augent.app/install.sh | bash"
        ) from err

    from .config import get_config

    cfg = get_config()
    audio_path = arguments.get("audio_path")
    model_size = arguments.get("model_size", cfg["model_size"])
    num_speakers = arguments.get("num_speakers")

    if not audio_path:
        raise ValueError("Missing required parameter: audio_path")

    result = identify_speakers(
        audio_path,
        model_size=model_size,
        num_speakers=num_speakers,
    )

    return {
        "speakers": result["speakers"],
        "segment_count": len(result["segments"]),
        "segments": result["segments"],
        "duration": result["duration"],
        "duration_formatted": result["duration_formatted"],
        "language": result["language"],
        "cached": result.get("cached", False),
        "model_used": model_size,
    }


def handle_search_memory(arguments: dict) -> dict:
    """Handle search_memory tool call."""
    from .config import get_config

    cfg = get_config()
    query = arguments.get("query")
    mode = arguments.get("mode", "keyword")
    top_k = arguments.get("top_k", 10)
    output = arguments.get("output")
    context_words = arguments.get("context_words", cfg["context_words"])
    dedup_seconds = arguments.get("dedup_seconds", 0)

    if not query:
        raise ValueError("Missing required parameter: query")

    if mode == "semantic":
        try:
            from .embeddings import search_memory
        except ImportError as err:
            raise RuntimeError(
                "Missing dependencies: sentence-transformers. "
                "Install with: pip install sentence-transformers"
            ) from err
        result = search_memory(
            query,
            top_k=top_k,
            mode="semantic",
            output=output,
            context_words=context_words,
            dedup_seconds=dedup_seconds,
        )
    else:
        from .embeddings import search_memory

        result = search_memory(query, top_k=top_k, mode="keyword", output=output)

    # Add YouTube timestamp links where source_url is YouTube
    for r in result.get("results", []):
        source_url = r.get("source_url", "")
        if source_url and _extract_youtube_id(source_url):
            secs = r.get("start", 0)
            yt_link = _youtube_timestamp_link(source_url, secs)
            if yt_link:
                r["youtube_link"] = yt_link

    return result


def handle_deep_search(arguments: dict) -> dict:
    """Handle deep_search tool call."""
    try:
        from .embeddings import deep_search
    except ImportError as err:
        raise RuntimeError(
            "Missing dependencies: sentence-transformers. "
            "Install with: pip install sentence-transformers"
        ) from err

    from .config import get_config

    cfg = get_config()
    audio_path = arguments.get("audio_path")
    query = arguments.get("query")
    model_size = arguments.get("model_size", cfg["model_size"])
    top_k = arguments.get("top_k", 5)
    output = arguments.get("output")
    context_words = arguments.get("context_words", cfg["context_words"])
    dedup_seconds = arguments.get("dedup_seconds", 0)
    clip = arguments.get("clip", False)
    clip_padding = arguments.get("clip_padding", cfg["clip_padding"])

    if not audio_path:
        raise ValueError("Missing required parameter: audio_path")
    if not query:
        raise ValueError("Missing required parameter: query")

    result = deep_search(
        audio_path,
        query,
        model_size=model_size,
        top_k=top_k,
        context_words=context_words,
        dedup_seconds=dedup_seconds,
    )

    # Add YouTube links if source is YouTube
    source_url = _downloaded_urls.get(os.path.abspath(audio_path), "")
    if not source_url:
        from .memory import get_transcription_memory

        mem = get_transcription_memory()
        source_url = mem.get_source_url(audio_path, model_size)
        if not source_url:
            source_url = mem.get_source_url_by_hash(audio_path)

    if source_url and _extract_youtube_id(source_url):
        for r in result.get("results", []):
            secs = r.get("start", 0)
            yt_link = _youtube_timestamp_link(source_url, secs)
            if yt_link:
                r["youtube_link"] = yt_link

    # Write output file if requested
    if output and result.get("results"):
        cols = ["timestamp", "text", "similarity"]
        if any(r.get("youtube_link") for r in result["results"]):
            cols.append("youtube_link")
        result["output_path"] = _write_output_file(
            output,
            result["results"],
            columns=cols,
            bold_columns=["timestamp"],
        )

    # Export clips around matches if requested
    if clip and source_url:
        ranges = [
            (float(r.get("start", 0)), float(r.get("end", r.get("start", 0))))
            for r in result.get("results", [])
            if r.get("start", 0)
        ]
        if ranges:
            result["clips"] = _export_clips_for_matches(
                source_url, time_ranges=ranges, padding=clip_padding
            )
        else:
            result["clips"] = []
            result["clip_note"] = "No matches with timestamps to clip."
    elif clip and not source_url:
        result["clips"] = []
        result["clip_note"] = (
            "No source URL found for this audio file. "
            "Clips can only be exported when the audio was downloaded from a URL."
        )

    return result


def handle_chapters(arguments: dict) -> dict:
    """Handle chapters tool call."""
    try:
        from .embeddings import detect_chapters
    except ImportError as err:
        raise RuntimeError(
            "Missing dependencies: sentence-transformers. "
            "Install with: pip install sentence-transformers"
        ) from err

    from .config import get_config

    cfg = get_config()
    audio_path = arguments.get("audio_path")
    model_size = arguments.get("model_size", cfg["model_size"])
    sensitivity = arguments.get("sensitivity", 0.4)

    if not audio_path:
        raise ValueError("Missing required parameter: audio_path")

    result = detect_chapters(
        audio_path,
        model_size=model_size,
        sensitivity=sensitivity,
    )

    # Trim chapter text to a short preview to avoid massive responses (~17k tokens)
    for chapter in result.get("chapters", []):
        text = chapter.get("text", "")
        words = text.split()
        if len(words) > 30:
            chapter["text"] = " ".join(words[:30]) + "..."

    return result


_tts_jobs = {}


def handle_text_to_speech(arguments: dict) -> dict:
    """Handle text_to_speech tool call. Runs in background subprocess, returns instantly."""
    import os
    import shutil
    import subprocess
    import tempfile
    import uuid

    # Check status of a running job
    job_id = arguments.get("job_id")
    if job_id:
        job = _tts_jobs.get(job_id)
        if not job:
            raise ValueError(f"Unknown job: {job_id}")
        proc = job["proc"]
        poll = proc.poll()
        if poll is None:
            return {
                "status": "generating",
                "job_id": job_id,
                "message": "TTS is still running. Check again in a few seconds.",
            }
        # Done — read result
        stdout = proc.stdout.read()
        proc.stdout.close()
        try:
            os.unlink(job["script"])
        except OSError:
            pass
        if poll != 0:
            del _tts_jobs[job_id]
            raise RuntimeError("TTS generation failed")
        result = json.loads(stdout.strip())
        if "error" in result:
            del _tts_jobs[job_id]
            raise RuntimeError(result["error"])
        del _tts_jobs[job_id]
        result["status"] = "complete"
        result["job_id"] = job_id
        return result

    from .config import get_config

    cfg = get_config()
    text = arguments.get("text")
    file_path = arguments.get("file_path")
    voice = arguments.get("voice", cfg["tts_voice"])
    output_dir = arguments.get("output_dir", cfg["notes_output_dir"])
    output_filename = arguments.get("output_filename")
    speed = arguments.get("speed", cfg["tts_speed"])

    if not text and not file_path:
        raise ValueError("Either text or file_path is required")

    # Build a Python script to run TTS in a completely separate process
    script = f"""
import json, sys, os
_real_stdout = os.dup(1)
sys.stdout = open('/dev/null', 'w')
sys.stderr = open('/dev/null', 'w')
os.dup2(os.open('/dev/null', os.O_WRONLY), 1)
os.dup2(os.open('/dev/null', os.O_WRONLY), 2)
from augent.tts import text_to_speech, read_aloud
try:
    if {repr(file_path)}:
        result = read_aloud({repr(file_path)}, voice={repr(voice)}, speed={speed})
    else:
        result = text_to_speech(
            text={repr(text)},
            voice={repr(voice)},
            output_dir={repr(output_dir)},
            output_filename={repr(output_filename)},
            speed={speed},
        )
    os.dup2(_real_stdout, 1)
    sys.stdout = os.fdopen(_real_stdout, 'w')
    print(json.dumps(result))
except Exception as e:
    os.dup2(_real_stdout, 1)
    sys.stdout = os.fdopen(_real_stdout, 'w')
    print(json.dumps({{"error": str(e)}}))
"""

    with tempfile.NamedTemporaryFile(mode="w", suffix=".py", delete=False) as f:
        f.write(script)
        script_path = f.name

    python_bin = shutil.which("python3") or sys.executable
    proc = subprocess.Popen(
        [python_bin, script_path],
        stdin=subprocess.DEVNULL,
        stdout=subprocess.PIPE,
        stderr=subprocess.DEVNULL,
        text=True,
    )

    job_id = str(uuid.uuid4())[:8]
    _tts_jobs[job_id] = {"proc": proc, "script": script_path}

    return {
        "status": "started",
        "job_id": job_id,
        "message": f"TTS generation started in background. Call text_to_speech with job_id='{job_id}' to check status.",
    }


def handle_separate_audio(arguments: dict) -> dict:
    """Handle separate_audio tool call."""
    audio_path = arguments.get("audio_path")
    vocals_only = arguments.get("vocals_only", True)
    model = arguments.get("model", "htdemucs")

    if not audio_path:
        raise ValueError("Missing required parameter: audio_path")

    audio_path = os.path.expanduser(audio_path)
    if not os.path.exists(audio_path):
        raise FileNotFoundError(f"Audio file not found: {audio_path}")

    from .separator import separate_audio

    two_stems = "vocals" if vocals_only else None

    result = separate_audio(
        audio_path,
        model=model,
        two_stems=two_stems,
    )

    response = {
        "stems": result["stems"],
        "model": result["model"],
        "source_file": result["source_file"],
        "cached": result["cached"],
        "output_dir": result["output_dir"],
    }

    # Highlight the vocals path for easy piping into transcribe_audio
    vocals_path = result["stems"].get("vocals")
    if vocals_path:
        response["vocals_path"] = vocals_path
        response["hint"] = (
            "Use the vocals_path as the audio_path in transcribe_audio, "
            "search_audio, deep_search, or any other tool for clean results."
        )

    return response


def _export_clips_for_matches(
    source_url: str,
    timestamps: list[float] | None = None,
    padding: int = 15,
    time_ranges: list[tuple[float, float]] | None = None,
) -> list[dict]:
    """Export video clips around match timestamps or time ranges.

    Accepts either:
      - timestamps: list of point-in-time matches (padding applied symmetrically)
      - time_ranges: list of (start, end) ranges (padding added before start and after end)

    Merges overlapping time ranges to avoid redundant downloads.
    Returns a list of clip info dicts.
    """
    # Build padded time ranges from either input format
    ranges = []

    if time_ranges:
        for seg_start, seg_end in time_ranges:
            # Add padding around the natural segment boundaries
            clip_start = max(0, seg_start - padding)
            clip_end = seg_end + padding
            ranges.append((clip_start, clip_end, seg_start))
    elif timestamps:
        for ts in sorted(set(timestamps)):
            start = max(0, ts - padding)
            end = ts + padding
            ranges.append((start, end, ts))
    else:
        return []

    # Sort by start time
    ranges.sort(key=lambda x: x[0])

    # Merge overlapping ranges
    merged: list[tuple[float, float, list[float]]] = []
    for start, end, ts in ranges:
        if merged and start <= merged[-1][1]:
            prev_start, prev_end, prev_ts_list = merged[-1]
            merged[-1] = (prev_start, max(prev_end, end), prev_ts_list + [ts])
        else:
            merged.append((start, end, [ts]))

    clips = []
    for start, end, ts_list in merged:
        try:
            clip_result = handle_clip_export(
                {
                    "url": source_url,
                    "start": start,
                    "end": end,
                }
            )
            clip_result["match_timestamps"] = ts_list
            clips.append(clip_result)
        except Exception as e:
            clips.append(
                {
                    "error": str(e),
                    "start": start,
                    "end": end,
                    "match_timestamps": ts_list,
                }
            )

    return clips


def handle_highlights(arguments: dict) -> dict:
    """Handle highlights tool call — export best moments from a transcription."""
    try:
        from .embeddings import deep_search, detect_chapters
    except ImportError as err:
        raise RuntimeError(
            "Missing dependencies: sentence-transformers. "
            "Install with: pip install sentence-transformers"
        ) from err

    from .config import get_config

    cfg = get_config()
    audio_path = arguments.get("audio_path")
    query = arguments.get("query")
    top_k = arguments.get("top_k", 5)
    model_size = arguments.get("model_size", cfg["model_size"])
    clip = arguments.get("clip", False)
    clip_padding = arguments.get("clip_padding", cfg["clip_padding"])
    context_words = arguments.get("context_words", 40)

    if not audio_path:
        raise ValueError("Missing required parameter: audio_path")

    highlights = []

    if query:
        # Focused mode: semantic search for the query
        search_result = deep_search(
            audio_path,
            query,
            model_size=model_size,
            top_k=top_k,
            context_words=context_words,
            dedup_seconds=30,
        )
        for r in search_result.get("results", []):
            highlights.append(
                {
                    "start": r["start"],
                    "end": r["end"],
                    "timestamp": r["timestamp"],
                    "text": r["text"],
                    "score": round(r["similarity"], 3),
                    "mode": "focused",
                }
            )
    else:
        # Auto mode: use chapters to find topic boundaries, then rank by density
        # Get chapters with moderate sensitivity for meaningful segments
        chapter_result = detect_chapters(
            audio_path,
            model_size=model_size,
            sensitivity=0.3,
        )
        chapters = chapter_result.get("chapters", [])

        if not chapters:
            raise ValueError("No chapters detected — audio may be too short or uniform")

        # Score each chapter by segment density (more segments = more content)
        # and prefer chapters that aren't too short or too long
        scored = []
        for ch in chapters:
            duration = ch["end"] - ch["start"]
            seg_count = ch.get("segment_count", 1)
            # Prefer medium-length segments (30s-120s) with high density
            if duration < 5:
                continue
            density = seg_count / max(duration, 1)
            # Penalize very short (<15s) and very long (>180s) chapters
            length_score = 1.0
            if duration < 15:
                length_score = 0.5
            elif duration > 180:
                length_score = 0.7
            score = density * length_score
            scored.append((score, ch))

        # Sort by score descending, take top_k
        scored.sort(key=lambda x: x[0], reverse=True)
        for score, ch in scored[:top_k]:
            # Get full text for the chapter via deep_search on a representative query
            text = ch.get("text", "")
            highlights.append(
                {
                    "start": ch["start"],
                    "end": ch["end"],
                    "timestamp": ch["start_timestamp"],
                    "text": text,
                    "score": round(score, 4),
                    "mode": "auto",
                    "chapter_number": ch["chapter_number"],
                    "duration": round(ch["end"] - ch["start"], 1),
                }
            )

        # Sort highlights chronologically
        highlights.sort(key=lambda x: x["start"])

    result = {
        "audio_path": audio_path,
        "mode": "focused" if query else "auto",
        "query": query,
        "highlight_count": len(highlights),
        "highlights": highlights,
        "model_used": model_size,
    }

    # Export clips if requested
    if clip:
        source_url = _downloaded_urls.get(os.path.abspath(audio_path), "")
        if not source_url:
            from .memory import get_transcription_memory

            mem = get_transcription_memory()
            source_url = mem.get_source_url(audio_path, model_size)
            if not source_url:
                source_url = mem.get_source_url_by_hash(audio_path)

        if source_url:
            ranges = [(h["start"], h["end"]) for h in highlights if "end" in h]
            if ranges:
                result["clips"] = _export_clips_for_matches(
                    source_url, time_ranges=ranges, padding=clip_padding
                )
            else:
                result["clips"] = []
                result["clip_note"] = "No highlights found to clip."
        else:
            result["clips"] = []
            result["clip_note"] = (
                "No source URL found for this audio file. "
                "Clips can only be exported when the audio was downloaded from a URL."
            )

    # Add YouTube links if available
    source_url = _downloaded_urls.get(os.path.abspath(audio_path), "")
    if not source_url:
        from .memory import get_transcription_memory

        mem = get_transcription_memory()
        source_url = mem.get_source_url(audio_path, model_size)
        if not source_url:
            source_url = mem.get_source_url_by_hash(audio_path)

    if source_url and _extract_youtube_id(source_url):
        for h in highlights:
            secs = h.get("start", 0)
            yt_link = _youtube_timestamp_link(source_url, secs)
            if yt_link:
                h["youtube_link"] = yt_link

    return result


def handle_tag(arguments: dict) -> dict:
    """Handle tag tool call — add, remove, or list tags on a transcription."""
    from .memory import get_transcription_memory

    cache_key = arguments.get("cache_key")
    action = arguments.get("action", "list")
    tags = arguments.get("tags", [])

    if not cache_key:
        return {"error": "cache_key is required"}

    mem = get_transcription_memory()

    if action == "add":
        if not tags:
            return {"error": "tags array is required for add action"}
        added = mem.add_tags(cache_key, tags, category="topic", source="auto")
        return {"action": "add", "cache_key": cache_key, "added": added}
    elif action == "remove":
        if not tags:
            return {"error": "tags array is required for remove action"}
        removed = mem.remove_tags(cache_key, tags)
        return {"action": "remove", "cache_key": cache_key, "removed": removed}
    elif action == "list":
        tag_list = mem.get_tags(cache_key)
        return {"action": "list", "cache_key": cache_key, "tags": tag_list}
    else:
        return {"error": f"Unknown action: {action}"}


def handle_rebuild_graph(arguments: dict) -> dict:
    """Handle rebuild_graph tool call — rebuild Obsidian graph view data."""
    from .graph import rebuild_graph
    from .memory import get_transcription_memory

    memory = get_transcription_memory()
    min_members = arguments.get("min_moc_members", 3)

    result = rebuild_graph(memory)

    # Re-run MOC generation with custom threshold if specified
    if min_members != 3:
        from .graph import generate_mocs

        moc_paths = generate_mocs(memory, min_members=min_members)
        result["mocs_generated"] = len(moc_paths)

    return {
        "success": True,
        "migration": result["migration"],
        "related_links_computed": result["related_computed"],
        "mocs_generated": result["mocs_generated"],
        "memory_dir": str(memory.md_dir),
        "hint": (
            "Point Obsidian at the memory_dir path as a vault (or add it to an existing vault). "
            "Enable Tags in graph view settings to see tag hub nodes. "
            "Use graph view Groups to color-code by type (transcription, notes, moc, translation)."
        ),
    }


def _get_obsidian_vault():
    """Get the first open Obsidian vault path from Obsidian's config."""
    from pathlib import Path

    config_path = (
        Path.home() / "Library" / "Application Support" / "obsidian" / "obsidian.json"
    )
    if not config_path.exists():
        return None
    try:
        with open(config_path) as f:
            data = json.loads(f.read())
        vaults = data.get("vaults", {})
        # Prefer the open vault, otherwise take the first one
        for v in vaults.values():
            if v.get("open"):
                p = Path(v["path"])
                if p.is_dir():
                    return p
        # No open vault, take any
        for v in vaults.values():
            p = Path(v["path"])
            if p.is_dir():
                return p
    except Exception:
        pass
    return None


def _score_visual_necessity(segments: list, segment_embeddings=None) -> list:
    """Score each transcript segment for visual necessity (0.0-1.0).

    Returns list of (segment_index, score, reason) tuples for ALL segments.
    Uses a hybrid approach: pattern matching + semantic similarity + heuristics.
    """
    import numpy as np

    # --- Pattern matching (fast, high precision) ---
    VISUAL_PATTERNS = [
        # Explicit visual references
        (
            _re.compile(
                r"\b(as you can see|you can see|you'll see|see here|shown here|look at)\b",
                _re.I,
            ),
            0.9,
            "explicit visual reference",
        ),
        (
            _re.compile(r"\b(on screen|on the screen|on your screen)\b", _re.I),
            0.9,
            "on-screen reference",
        ),
        (
            _re.compile(
                r"\b(here'?s what it looks like|this is what it looks like|looks like this)\b",
                _re.I,
            ),
            0.9,
            "visual demonstration",
        ),
        # UI actions
        (
            _re.compile(
                r"\b(click|tap|press|hit|select|toggle|check|uncheck)\s+(on|the|this|that|here)\b",
                _re.I,
            ),
            0.85,
            "UI action",
        ),
        (
            _re.compile(r"\b(drag|drop|swipe|scroll|hover)\b", _re.I),
            0.85,
            "spatial UI action",
        ),
        (
            _re.compile(
                r"\b(navigate to|go to|open up|expand|collapse|minimize|maximize)\b",
                _re.I,
            ),
            0.75,
            "navigation action",
        ),
        # UI elements
        (
            _re.compile(
                r"\b(button|icon|menu|dropdown|sidebar|toolbar|popup|modal|dialog|tooltip)\b",
                _re.I,
            ),
            0.7,
            "UI element reference",
        ),
        (
            _re.compile(
                r"\b(dashboard|chart|graph|table|spreadsheet|form|field|checkbox|slider)\b",
                _re.I,
            ),
            0.7,
            "data visualization",
        ),
        # Spatial/positional
        (
            _re.compile(
                r"\b(top right|top left|bottom right|bottom left|upper right|upper left|lower right|lower left)\b",
                _re.I,
            ),
            0.8,
            "spatial position",
        ),
        (
            _re.compile(
                r"\b(over here|right here|right there|over there|this area|this section|this part)\b",
                _re.I,
            ),
            0.85,
            "deictic reference",
        ),
        # Demonstration language
        (
            _re.compile(
                r"\b(let me show|i'll show|i'm going to show|watch what happens|watch this|notice how)\b",
                _re.I,
            ),
            0.85,
            "demonstration",
        ),
        (
            _re.compile(
                r"\b(step \d|first.{0,20}(click|select|open|type|enter)|next.{0,20}(click|select|open|type))\b",
                _re.I,
            ),
            0.8,
            "step-by-step instruction",
        ),
        # Code/terminal
        (
            _re.compile(
                r"\b(the (code|output|error|terminal|console) (shows|says|reads|displays))\b",
                _re.I,
            ),
            0.75,
            "code/terminal output",
        ),
        (
            _re.compile(
                r"\b(type|enter|run|execute|paste)\s+(this|the command|the following|in the)\b",
                _re.I,
            ),
            0.7,
            "command input",
        ),
        # Screen recording cues
        (
            _re.compile(
                r"\b(recording|screen share|screen cast|let me walk you through)\b",
                _re.I,
            ),
            0.7,
            "screen recording",
        ),
        (
            _re.compile(r"\b(you('ll| will) (notice|see|find))\b", _re.I),
            0.75,
            "visual callout",
        ),
    ]

    # --- Semantic anchors (for embedding-based scoring) ---
    VISUAL_ANCHORS = [
        "demonstrating a user interface action on screen",
        "showing what something looks like visually",
        "clicking buttons and navigating menus in software",
        "pointing at something specific on a screen or dashboard",
        "step by step tutorial showing visual actions",
        "presenting a chart, graph, or data visualization",
        "showing code output or terminal results on screen",
        "walking through a workflow on screen",
    ]
    NON_VISUAL_ANCHORS = [
        "expressing an opinion or abstract thought",
        "discussing general concepts and ideas",
        "telling a personal story or anecdote",
        "greeting the audience or introducing the topic",
        "summarizing what was previously discussed",
        "background music or sound effects playing",
    ]

    # Score each segment
    scored = []

    # Pre-compute semantic scores if embeddings available
    semantic_scores = None
    if segment_embeddings is not None and len(segment_embeddings) > 0:
        try:
            from .embeddings import _cosine_similarity, _get_embedding_model_cache

            model = _get_embedding_model_cache().get()
            visual_embs = model.encode(
                VISUAL_ANCHORS, convert_to_numpy=True, show_progress_bar=False
            )
            non_visual_embs = model.encode(
                NON_VISUAL_ANCHORS, convert_to_numpy=True, show_progress_bar=False
            )

            semantic_scores = []
            for i in range(len(segment_embeddings)):
                seg_emb = segment_embeddings[i].reshape(1, -1)
                vis_sims = _cosine_similarity(seg_emb, visual_embs)
                non_vis_sims = _cosine_similarity(seg_emb, non_visual_embs)
                vis_max = float(np.max(vis_sims))
                non_vis_max = float(np.max(non_vis_sims))
                # Net visual score: how much more visual than non-visual
                semantic_scores.append(
                    max(0.0, min(1.0, (vis_max - non_vis_max + 0.3) / 0.6))
                )
        except Exception:
            semantic_scores = None

    for i, seg in enumerate(segments):
        text = seg.get("text", "").strip()

        # Pattern matching: take best match
        best_pattern_score = 0.0
        best_reason = ""
        for pattern, score, reason in VISUAL_PATTERNS:
            if pattern.search(text):
                if score > best_pattern_score:
                    best_pattern_score = score
                    best_reason = reason

        # Semantic score
        sem_score = (
            semantic_scores[i] if semantic_scores and i < len(semantic_scores) else 0.0
        )

        # Heuristic adjustments
        heuristic_mult = 1.0
        words = text.split()
        if len(words) < 3:
            heuristic_mult *= 0.4  # Very short segments are likely filler
        if text.endswith("?") and best_pattern_score == 0:
            heuristic_mult *= 0.5  # Questions without visual keywords

        # Combine scores
        if best_pattern_score >= 0.8:
            combined = best_pattern_score
            reason = best_reason
        elif best_pattern_score > 0:
            combined = 0.6 * best_pattern_score + 0.4 * sem_score
            reason = best_reason
        else:
            combined = sem_score * 0.7
            reason = "semantic: visual content detected" if sem_score > 0.5 else ""

        final_score = min(1.0, combined * heuristic_mult)
        scored.append((i, final_score, reason))

    # Boost sequential visual segments (tutorial sequences)
    for i in range(len(scored)):
        if i >= 2:
            _, s1, _ = scored[i - 2]
            _, s2, _ = scored[i - 1]
            idx, s3, reason = scored[i]
            if s1 > 0.4 and s2 > 0.4 and s3 > 0.3:
                scored[i] = (idx, min(1.0, s3 * 1.25), reason or "tutorial sequence")

    # Suppress intro B-roll false positives:
    # First 30 seconds require score > 0.9 to qualify.
    # Detect intro pattern: 3+ qualifying segments within first 15 seconds = B-roll burst, suppress all.
    intro_cutoff = 30.0
    intro_burst_cutoff = 15.0
    intro_burst_count = 0
    for seg_idx, score, _reason in scored:
        seg_start = segments[seg_idx].get("start", 0)
        if seg_start < intro_burst_cutoff and score > 0.4:
            intro_burst_count += 1
    intro_is_broll = intro_burst_count >= 3

    if intro_is_broll:
        for i, (seg_idx, score, reason) in enumerate(scored):
            seg_start = segments[seg_idx].get("start", 0)
            if seg_start < intro_cutoff and score < 0.95:
                scored[i] = (seg_idx, score * 0.3, reason)  # Heavily suppress
    else:
        # Even without burst detection, require higher confidence in first 30s
        for i, (seg_idx, score, reason) in enumerate(scored):
            seg_start = segments[seg_idx].get("start", 0)
            if seg_start < intro_cutoff and score < 0.9:
                scored[i] = (seg_idx, score * 0.6, reason)

    return scored


def _extract_best_frame(
    video_path: str, ts: float, duration: float, out_path: str
) -> bool:
    """Extract the best frame from 3 candidates around a timestamp.

    Takes frames at ts, ts+1s, ts+2s and picks the one with the most
    visual information (highest edge density = most UI/text content).
    Returns True if a frame was saved to out_path.
    """
    import tempfile

    candidates = []
    offsets = [0.0, 1.0, 2.0]

    for offset in offsets:
        t = ts + offset
        if t >= duration - 0.1:
            continue

        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        tmp.close()

        cmd = [
            "ffmpeg",
            "-y",
            "-ss",
            str(t),
            "-i",
            video_path,
            "-frames:v",
            "1",
            "-vf",
            "scale='min(1280,iw)':-1",
            "-q:v",
            "2",
            tmp.name,
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
        if result.returncode != 0 or not os.path.exists(tmp.name):
            try:
                os.unlink(tmp.name)
            except OSError:
                pass
            continue

        # Score by edge density: convert to grayscale, compute gradient magnitude
        try:
            from PIL import Image

            img = Image.open(tmp.name).convert("L").resize((320, 180))
            import numpy as np

            arr = np.array(img, dtype=np.float32)
            # Sobel-like gradient: horizontal + vertical differences
            gx = np.abs(arr[:, 1:] - arr[:, :-1])
            gy = np.abs(arr[1:, :] - arr[:-1, :])
            edge_score = float(np.mean(gx) + np.mean(gy))
            candidates.append((tmp.name, edge_score))
        except Exception:
            # If PIL not available, just use the first candidate
            candidates.append((tmp.name, 0.0))

    if not candidates:
        return False

    # Pick highest edge score
    candidates.sort(key=lambda x: x[1], reverse=True)
    best_path = candidates[0][0]

    # Move best to output, clean up others
    import shutil

    shutil.move(best_path, out_path)
    for path, _ in candidates:
        if path != best_path:
            try:
                os.unlink(path)
            except OSError:
                pass

    return True


def _dedup_frames(frame_info: list) -> list:
    """Remove near-duplicate frames using average perceptual hashing.

    Downscales each frame to 8x8 grayscale, computes a 64-bit hash,
    compares hamming distance. If two frames have distance < 5, drops
    the lower-scored one. No new dependencies (PIL + numpy).
    """
    if len(frame_info) <= 1:
        return frame_info

    try:
        import numpy as np
        from PIL import Image
    except ImportError:
        return frame_info  # Can't dedup without PIL

    def _ahash(path: str) -> int:
        """Compute average hash: 8x8 grayscale, compare to mean."""
        img = Image.open(path).convert("L").resize((8, 8))
        arr = np.array(img, dtype=np.float32)
        mean = arr.mean()
        bits = (arr > mean).flatten()
        h = 0
        for bit in bits:
            h = (h << 1) | int(bit)
        return h

    def _hamming(a: int, b: int) -> int:
        return bin(a ^ b).count("1")

    # Compute hashes
    hashes = []
    for fi in frame_info:
        try:
            hashes.append(_ahash(fi["path"]))
        except Exception:
            hashes.append(None)

    # Mark duplicates (keep higher-scored frame)
    keep = [True] * len(frame_info)
    for i in range(len(frame_info)):
        if not keep[i] or hashes[i] is None:
            continue
        for j in range(i + 1, len(frame_info)):
            if not keep[j] or hashes[j] is None:
                continue
            if _hamming(hashes[i], hashes[j]) < 5:
                # Drop the lower-scored one
                score_i = frame_info[i].get("score", 0)
                score_j = frame_info[j].get("score", 0)
                if score_i >= score_j:
                    keep[j] = False
                    # Clean up the file
                    try:
                        os.unlink(frame_info[j]["path"])
                    except OSError:
                        pass
                else:
                    keep[i] = False
                    try:
                        os.unlink(frame_info[i]["path"])
                    except OSError:
                        pass
                    break  # i is dropped, move on

    return [fi for fi, k in zip(frame_info, keep, strict=False) if k]


def handle_visual(arguments: dict) -> dict:
    """Handle visual tool call — extract frames at moments that matter."""
    import shutil
    from pathlib import Path

    from .config import get_config
    from .core import transcribe_audio
    from .embeddings import (
        _cosine_similarity,
        _get_embedding_model_cache,
        _get_or_compute_embeddings,
        _ranked_semantic_search,
    )
    from .memory import get_transcription_memory

    cfg = get_config()

    video_path = arguments.get("video_path")
    url = arguments.get("url")

    if not video_path and not url:
        raise ValueError(
            "Provide video_path (local file) or url (downloads video automatically)."
        )

    if url and not video_path:
        # Download video from URL — prefer brew yt-dlp (stays current, avoids pip SSL issues)
        ytdlp = shutil.which(
            "yt-dlp", path="/opt/homebrew/bin:/usr/local/bin"
        ) or shutil.which("yt-dlp")
        if not ytdlp:
            raise RuntimeError("yt-dlp not found. Install with: brew install yt-dlp")
        download_dir = os.path.expanduser("~/Downloads")
        cmd = [
            ytdlp,
            "-f",
            "bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best",
            "--merge-output-format",
            "mp4",
            "--no-playlist",
            "-o",
            os.path.join(download_dir, "%(title)s.%(ext)s"),
            "--print",
            "after_move:filepath",
            url,
        ]
        dl_result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
        if dl_result.returncode != 0:
            raise RuntimeError(
                f"Video download failed: {dl_result.stderr.strip()[:300]}"
            )
        video_path = dl_result.stdout.strip().split("\n")[-1]
        if not os.path.exists(video_path):
            raise RuntimeError(
                f"Video download completed but file not found: {video_path}"
            )

    video_path = os.path.expanduser(video_path)
    if not os.path.exists(video_path):
        raise FileNotFoundError(f"Video file not found: {video_path}")

    if not shutil.which("ffmpeg") or not shutil.which("ffprobe"):
        raise RuntimeError(
            "ffmpeg and ffprobe are required. Install with: brew install ffmpeg"
        )

    query = arguments.get("query")
    timestamps = arguments.get("timestamps")
    auto_mode = arguments.get("auto", False)
    assist_mode = arguments.get("assist", False)
    clear = arguments.get("clear", False)
    model_size = arguments.get("model_size", cfg.get("model_size", "tiny"))
    max_frames = int(
        arguments.get("max_frames", cfg.get("visual_context_max_frames", 30))
    )
    top_k = int(arguments.get("top_k", 10))
    context_words = int(arguments.get("context_words", 40))

    if max_frames < 1:
        raise ValueError("max_frames must be at least 1")

    # Handle clear: remove all frames and visual context .md for this video
    video_stem_for_clear = os.path.splitext(os.path.basename(video_path))[0]
    safe_name_for_clear = _re.sub(r"[^\w\s-]", "", video_stem_for_clear).strip()[:80]
    if clear:
        vault_dir = _get_obsidian_vault()
        removed_frames = 0
        removed_md = False
        # Remove frames from vault
        if vault_dir:
            frames_dir = vault_dir / "External Files" / "visual" / safe_name_for_clear
            if frames_dir.is_dir():
                for f in frames_dir.glob("*.png"):
                    f.unlink()
                    removed_frames += 1
                try:
                    frames_dir.rmdir()
                except OSError:
                    pass
        # Remove visual context .md from Desktop
        md_name = f"{safe_name_for_clear} - Visual Context.md"
        md_desktop = (
            Path(os.path.expanduser(cfg.get("notes_output_dir", "~/Desktop"))) / md_name
        )
        if md_desktop.exists():
            md_desktop.unlink()
            removed_md = True
        # Also remove from vault External Files if hard-linked
        if vault_dir:
            md_vault = vault_dir / "External Files" / md_name
            if md_vault.exists():
                md_vault.unlink()
                removed_md = True

        if not query and not timestamps and not auto_mode:
            return {
                "cleared": True,
                "removed_frames": removed_frames,
                "removed_md": removed_md,
                "video_path": video_path,
            }

    # Determine mode
    if timestamps:
        mode = "manual"
    elif query:
        mode = "query"
    elif auto_mode:
        mode = "auto"
    elif assist_mode:
        mode = "assist"
    else:
        raise ValueError(
            "Provide one of: query (describe what you need visual context for), "
            "timestamps (list of seconds), auto: true (autonomous detection), "
            "or assist: true (flag visual gaps for manual screenshots)."
        )

    # Probe video duration
    ffprobe_cmd = [
        "ffprobe",
        "-v",
        "error",
        "-show_entries",
        "format=duration",
        "-of",
        "default=noprint_wrappers=1:nokey=1",
        video_path,
    ]
    probe_result = subprocess.run(
        ffprobe_cmd, capture_output=True, text=True, timeout=30
    )
    if probe_result.returncode != 0:
        raise RuntimeError(f"ffprobe failed: {probe_result.stderr.strip()[:200]}")
    video_duration = float(probe_result.stdout.strip())

    def _fmt_ts(seconds: float) -> str:
        m, s = divmod(int(seconds), 60)
        h, m = divmod(m, 60)
        if h > 0:
            return f"{h}:{m:02d}:{s:02d}"
        return f"{m}:{s:02d}"

    # Transcribe (reuses cache)
    transcription = transcribe_audio(video_path, model_size)
    segments = transcription["segments"]

    if not segments and mode != "manual":
        return {
            "video_path": video_path,
            "mode": mode,
            "frame_count": 0,
            "analyzed_segments": 0,
            "video_duration": round(video_duration, 1),
            "video_duration_formatted": _fmt_ts(video_duration),
            "frames": [],
            "hint": "No transcript segments found. The video may have no speech. Try manual mode with timestamps.",
        }

    # Memory setup
    memory = get_transcription_memory()
    audio_hash = memory.hash_audio_file(video_path)

    # Build extraction targets: list of (timestamp, score, reason, transcript_context)
    targets = []

    if mode == "manual":
        for ts in timestamps[:max_frames]:
            ts = float(ts)
            # Find nearest segment for context
            context = ""
            if segments:
                nearest = min(segments, key=lambda s: abs(s.get("start", 0) - ts))
                seg_idx = segments.index(nearest)
                parts = []
                for j in range(max(0, seg_idx - 1), min(len(segments), seg_idx + 2)):
                    parts.append(segments[j].get("text", "").strip())
                context = " ".join(parts)
            targets.append((ts, 1.0, "manual", context))

    elif mode == "query":
        # Use deep_search infrastructure to find matching moments
        segment_embeddings = _get_or_compute_embeddings(segments, audio_hash)

        embed_model = _get_embedding_model_cache().get()
        query_embedding = embed_model.encode(
            query, convert_to_numpy=True, show_progress_bar=False
        )

        segments_meta = [
            {"seg": seg, "seg_idx": i, "file_segments": segments}
            for i, seg in enumerate(segments)
        ]

        results = _ranked_semantic_search(
            query_embedding,
            segment_embeddings,
            segments_meta,
            query,
            top_k,
            context_words,
            dedup_seconds=5.0,
        )

        for r in results[:max_frames]:
            ts = r["start"] + 1.0  # Offset into the visual moment
            targets.append(
                (ts, round(r["similarity"], 3), f"query match: {query}", r["text"])
            )

    elif mode == "auto":
        # Autonomous detection using pattern + semantic scoring
        segment_embeddings = None
        try:
            segment_embeddings = _get_or_compute_embeddings(segments, audio_hash)
        except Exception:
            pass

        scored = _score_visual_necessity(segments, segment_embeddings)

        # Filter, sort by score, cap
        qualifying = [
            (idx, score, reason) for idx, score, reason in scored if score >= 0.4
        ]
        qualifying.sort(key=lambda x: x[1], reverse=True)
        qualifying = qualifying[:max_frames]
        qualifying.sort(key=lambda x: segments[x[0]].get("start", 0))

        # Deduplicate within 3 seconds
        deduped = []
        for idx, score, reason in qualifying:
            ts = segments[idx].get("start", 0)
            if deduped:
                last_ts = segments[deduped[-1][0]].get("start", 0)
                if abs(ts - last_ts) < 3.0:
                    if score > deduped[-1][1]:
                        deduped[-1] = (idx, score, reason)
                    continue
            deduped.append((idx, score, reason))

        for idx, score, reason in deduped:
            seg = segments[idx]
            ts = seg.get("start", 0) + 1.0
            parts = []
            for j in range(max(0, idx - 1), min(len(segments), idx + 2)):
                parts.append(segments[j].get("text", "").strip())
            targets.append((ts, score, reason, " ".join(parts)))

    elif mode == "assist":
        # Assist mode: detect visual gaps and cluster into time ranges.
        # Uses the same scoring pipeline as auto mode but returns gap
        # analysis instead of extracting frames.
        segment_embeddings = None
        try:
            segment_embeddings = _get_or_compute_embeddings(segments, audio_hash)
        except Exception:
            pass

        scored = _score_visual_necessity(segments, segment_embeddings)

        # Filter to qualifying segments (score >= 0.4)
        qualifying = [
            (idx, score, reason) for idx, score, reason in scored if score >= 0.4
        ]
        qualifying.sort(key=lambda x: segments[x[0]].get("start", 0))

        # Cluster consecutive qualifying segments into time ranges.
        # Segments within 15 seconds of each other merge into one gap.
        gaps = []
        for idx, score, reason in qualifying:
            seg = segments[idx]
            seg_start = seg.get("start", 0)
            seg_end = seg.get("end", seg_start + 3.0)

            if gaps and seg_start - gaps[-1]["end"] <= 15.0:
                # Extend current gap
                gap = gaps[-1]
                gap["end"] = seg_end
                gap["segments"].append(idx)
                gap["peak_score"] = max(gap["peak_score"], score)
                # Collect all unique reasons
                if reason and reason not in gap["reasons"]:
                    gap["reasons"].append(reason)
            else:
                # Start new gap
                gaps.append(
                    {
                        "start": seg_start,
                        "end": seg_end,
                        "segments": [idx],
                        "peak_score": score,
                        "reasons": [reason] if reason else [],
                    }
                )

        # Build structured output for each gap
        visual_gaps = []
        for i, gap in enumerate(gaps):
            # Collect transcript text across all segments in this gap
            transcript_parts = []
            for seg_idx in gap["segments"]:
                text = segments[seg_idx].get("text", "").strip()
                if text:
                    transcript_parts.append(text)
            transcript = " ".join(transcript_parts)

            # Generate a descriptive label from the reasons
            reasons = gap["reasons"]
            if any(
                r in ("UI action", "spatial UI action", "navigation action")
                for r in reasons
            ):
                screenshot_type = "UI interaction or navigation being described"
            elif any(r in ("data visualization",) for r in reasons):
                screenshot_type = "dashboard, chart, or data view being referenced"
            elif any(r in ("code/terminal output", "command input") for r in reasons):
                screenshot_type = "code, terminal, or command output being described"
            elif any(r in ("step-by-step instruction",) for r in reasons):
                screenshot_type = "step-by-step process being walked through"
            elif any(r in ("demonstration", "screen recording") for r in reasons):
                screenshot_type = "workflow or demonstration being shown"
            elif any(
                r
                in (
                    "explicit visual reference",
                    "on-screen reference",
                    "visual demonstration",
                )
                for r in reasons
            ):
                screenshot_type = "specific screen or visual the speaker is referencing"
            elif any(r in ("deictic reference", "spatial position") for r in reasons):
                screenshot_type = "specific UI element or area being pointed to"
            elif any("semantic" in r for r in reasons):
                screenshot_type = "visual context implied by the discussion"
            else:
                screenshot_type = "visual context for what the speaker is describing"

            visual_gaps.append(
                {
                    "gap_number": i + 1,
                    "start": round(gap["start"], 1),
                    "end": round(gap["end"], 1),
                    "start_formatted": _fmt_ts(gap["start"]),
                    "end_formatted": _fmt_ts(gap["end"]),
                    "duration_seconds": round(gap["end"] - gap["start"], 1),
                    "peak_score": round(gap["peak_score"], 2),
                    "screenshot_type": screenshot_type,
                    "reasons": reasons,
                    "transcript": transcript[:500],
                }
            )

        return {
            "video_path": video_path,
            "mode": "assist",
            "gap_count": len(visual_gaps),
            "analyzed_segments": len(segments),
            "video_duration": round(video_duration, 1),
            "video_duration_formatted": _fmt_ts(video_duration),
            "gaps": visual_gaps,
            "hint": (
                "These are moments where the speaker describes something visual "
                "but the video may not show it. For better results replicating "
                "their workflow, provide your own screenshots for these time ranges. "
                "Once you have screenshots, use visual() with timestamps to place them, "
                "or drop them directly into the Obsidian vault."
            ),
        }

    # Clamp timestamps to video bounds
    targets = [
        (min(max(ts, 0.1), video_duration - 0.1), score, reason, ctx)
        for ts, score, reason, ctx in targets
    ]

    # Sort chronologically
    targets.sort(key=lambda t: t[0])

    # Frame storage
    video_stem = os.path.splitext(os.path.basename(video_path))[0]
    safe_name = _re.sub(r"[^\w\s-]", "", video_stem).strip()[:80]
    # Short prefix for frame filenames (unique per video, readable in Obsidian)
    name_prefix = _re.sub(r"[^\w]", "_", safe_name).strip("_")[:40].lower().rstrip("_")

    # Frames go directly into the Obsidian vault so ![[]] embeds resolve.
    # The .md goes to Desktop — augent-obsidian hard-links it into the vault.
    # PNGs can't be hard-linked (only .md/.txt), so we write them into the vault directly.
    vault_dir = _get_obsidian_vault()

    if vault_dir:
        desktop_dir = str(vault_dir / "External Files" / "visual" / safe_name)
    else:
        desktop_dir = os.path.join(
            os.path.expanduser(cfg.get("notes_output_dir", "~/Desktop")),
            "visual",
            safe_name,
        )
    os.makedirs(desktop_dir, exist_ok=True)

    # Extract frames — picks best of 3 candidates per timestamp by edge density
    frame_info = []
    for _frame_num, (ts, score, reason, context) in enumerate(targets):
        # Unique filename: videoname_07m19s.png
        m, s = divmod(int(ts), 60)
        fname = f"{name_prefix}_{m:02d}m{s:02d}s.png"
        desktop_path = os.path.join(desktop_dir, fname)

        if not _extract_best_frame(video_path, ts, video_duration, desktop_path):
            continue

        frame_info.append(
            {
                "path": desktop_path,
                "filename": fname,
                "timestamp": round(ts, 1),
                "timestamp_formatted": _fmt_ts(ts),
                "score": round(score, 2) if isinstance(score, float) else score,
                "reason": reason,
                "transcript": context,
            }
        )

    # Remove near-duplicate frames (same UI state, minor cursor movement)
    frame_info = _dedup_frames(frame_info)

    # Create visual context .md on Desktop with frame embeds.
    # augent-obsidian hard-links it into the vault automatically.
    md_path = None
    if frame_info:
        try:
            md_name = f"{safe_name} - Visual Context.md"
            md_path = os.path.join(
                os.path.expanduser(cfg.get("notes_output_dir", "~/Desktop")),
                md_name,
            )

            md_lines = [
                f"# {video_stem} — Visual Context",
                "",
                f"**Query:** {query}" if query else "**Mode:** auto",
                f"**Duration:** {_fmt_ts(video_duration)}",
                f"**Frames:** {len(frame_info)}",
                "",
                "---",
                "",
            ]

            for fi in frame_info:
                md_lines.append(f"### {fi['timestamp_formatted']}")
                md_lines.append("")
                md_lines.append(f"![[{fi['filename']}]]")
                md_lines.append("")
                if fi.get("transcript"):
                    clean_transcript = _re.sub(
                        r"\*\*(.+?)\*\*", r"\1", fi["transcript"]
                    )
                    md_lines.append(f"> {clean_transcript[:300]}")
                    md_lines.append("")

            with open(md_path, "w", encoding="utf-8") as f:
                f.write("\n".join(md_lines))
        except Exception:
            md_path = None

    return {
        "video_path": video_path,
        "mode": mode,
        "query": query if mode == "query" else None,
        "frame_count": len(frame_info),
        "analyzed_segments": len(segments),
        "video_duration": round(video_duration, 1),
        "video_duration_formatted": _fmt_ts(video_duration),
        "frames_dir": desktop_dir,
        "md_path": md_path,
        "frames": frame_info,
        "hint": "Use the Read tool to view any frame PNG. Each frame includes the transcript of what was being said at that moment.",
    }


def handle_clip_export(arguments: dict) -> dict:
    """Handle clip_export tool call — download a video segment from a URL."""
    from .config import get_config

    cfg = get_config()
    url = arguments.get("url")
    start = arguments.get("start")
    end = arguments.get("end")
    output_dir = arguments.get(
        "output_dir", os.path.expanduser(cfg["notes_output_dir"])
    )
    output_filename = arguments.get("output_filename")

    if not url:
        raise ValueError("Missing required parameter: url")
    if start is None or end is None:
        raise ValueError("Missing required parameters: start and end")
    if end <= start:
        raise ValueError("end must be greater than start")

    output_dir = os.path.expanduser(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    ytdlp = shutil.which(
        "yt-dlp", path="/opt/homebrew/bin:/usr/local/bin"
    ) or shutil.which("yt-dlp")
    if not ytdlp:
        raise FileNotFoundError("yt-dlp not found. Install with: pip install yt-dlp")

    # Format times for yt-dlp --download-sections
    def fmt_time(s):
        m, sec = divmod(int(s), 60)
        h, m = divmod(m, 60)
        return f"{h:02d}:{m:02d}:{sec:02d}"

    section = f"*{fmt_time(start)}-{fmt_time(end)}"

    # Build output template
    if output_filename:
        out_template = os.path.join(output_dir, f"{output_filename}.%(ext)s")
    else:
        out_template = os.path.join(
            output_dir, "%(title)s_clip_%(section_start)s-%(section_end)s.%(ext)s"
        )

    cmd = [
        ytdlp,
        "--download-sections",
        section,
        "--force-keyframes-at-cuts",
        "--force-overwrites",
        "-f",
        "bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best",
        "--merge-output-format",
        "mp4",
        "--no-playlist",
        "-o",
        out_template,
        "--print",
        "after_move:filepath",
        url,
    ]

    result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)

    if result.returncode != 0:
        error_msg = result.stderr.strip()[-300:] if result.stderr else "Unknown error"
        raise RuntimeError(f"yt-dlp clip export failed: {error_msg}")

    output_lines = result.stdout.strip().split("\n")
    clip_path = output_lines[-1] if output_lines else None

    if not clip_path or not os.path.exists(clip_path):
        raise RuntimeError("Clip file not found after export")

    file_size = os.path.getsize(clip_path)
    duration = end - start

    return {
        "clip_path": clip_path,
        "url": url,
        "start": start,
        "end": end,
        "start_formatted": fmt_time(start),
        "end_formatted": fmt_time(end),
        "duration": duration,
        "duration_formatted": f"{int(duration // 60)}:{int(duration % 60):02d}",
        "file_size_mb": round(file_size / (1024 * 1024), 2),
    }


def _normalize_twitter_space_url(url: str) -> str:
    """Normalize Twitter/X Space URLs for compatibility."""
    url = url.replace("https://x.com/", "https://twitter.com/")
    if url.endswith("/peek"):
        url = url[:-5]
    return url


def _get_twitter_cookies_path() -> str:
    """Get path to Twitter cookies file, generating it from auth.json if needed."""
    augent_dir = os.path.expanduser("~/.augent")
    auth_path = os.path.join(augent_dir, "auth.json")
    cookies_path = os.path.join(augent_dir, "twitter_cookies.txt")

    if os.path.exists(auth_path):
        auth_mtime = os.path.getmtime(auth_path)
        cookies_mtime = os.path.getmtime(cookies_path) if os.path.exists(cookies_path) else 0

        if auth_mtime > cookies_mtime:
            with open(auth_path) as f:
                auth = json.load(f)
            auth_token = auth.get("auth_token", "")
            ct0 = auth.get("ct0", "")
            lines = [
                "# Netscape HTTP Cookie File",
                f".twitter.com\tTRUE\t/\tTRUE\t0\tauth_token\t{auth_token}",
                f".twitter.com\tTRUE\t/\tTRUE\t0\tct0\t{ct0}",
            ]
            os.makedirs(augent_dir, exist_ok=True)
            with open(cookies_path, "w") as f:
                f.write("\n".join(lines) + "\n")

    if os.path.exists(cookies_path) and os.path.getsize(cookies_path) > 0:
        return cookies_path

    return None


_SPACES_SETUP_INSTRUCTIONS = (
    "Twitter requires a one-time setup to download Spaces.\n\n"
    "Steps (30 seconds):\n"
    "1. Open Chrome > go to twitter.com (make sure you're logged in)\n"
    "2. Press F12 (or Cmd+Option+I) to open DevTools\n"
    "3. Click Application tab > Cookies > https://twitter.com\n"
    "4. Find auth_token — copy its Value\n"
    "5. Find ct0 — copy its Value\n"
    "6. Create the file ~/.augent/auth.json with:\n"
    '   {"auth_token": "PASTE_HERE", "ct0": "PASTE_HERE"}\n\n'
    "Your tokens are stored locally and only sent to Twitter's own servers to fetch audio. "
    "Augent never posts, DMs, follows, or modifies anything on your account. "
    "To revoke access anytime, simply log out of Twitter or delete ~/.augent/auth.json."
)


def handle_spaces(arguments: dict) -> dict:
    """Handle spaces tool call. Routes to download, check, or stop based on params."""
    recording_id = arguments.get("recording_id")
    stop = arguments.get("stop", False)

    if recording_id and stop:
        return _spaces_stop(arguments)
    elif recording_id:
        return _spaces_check(arguments)
    elif arguments.get("url"):
        return _spaces_download(arguments)
    else:
        raise ValueError("Provide either url (to start download) or recording_id (to check/stop)")


def _spaces_download(arguments: dict) -> dict:
    """Start a Twitter Space download. Auto-detects live vs ended."""
    import glob as glob_module

    url = arguments.get("url")
    output_dir = arguments.get("output_dir", os.path.expanduser("~/Downloads"))

    if not url:
        raise ValueError("Missing required parameter: url")

    cookies_path = _get_twitter_cookies_path()
    if not cookies_path:
        raise FileNotFoundError(_SPACES_SETUP_INSTRUCTIONS)

    os.makedirs(output_dir, exist_ok=True)
    url = _normalize_twitter_space_url(url)

    meta_cmd = [
        "yt-dlp",
        "--cookies", cookies_path,
        "--add-header", "Referer:https://twitter.com/",
        "--no-playlist", "--dump-json",
        url,
    ]
    meta_result = subprocess.run(meta_cmd, capture_output=True, text=True, timeout=30)

    if meta_result.returncode != 0:
        error = meta_result.stderr.strip()
        raise RuntimeError(f"Failed to fetch space info: {error[:300]}")

    meta = json.loads(meta_result.stdout)
    title = meta.get("title", "twitter_space")
    is_live = meta.get("is_live", False)

    before = set(glob_module.glob(os.path.join(output_dir, "*")))

    if is_live:
        stream_cmd = [
            "yt-dlp",
            "--cookies", cookies_path,
            "--add-header", "Referer:https://twitter.com/",
            "--no-playlist", "-g", "-f", "bestaudio",
            url,
        ]
        stream_result = subprocess.run(stream_cmd, capture_output=True, text=True, timeout=15)
        if stream_result.returncode != 0:
            raise RuntimeError("Failed to get stream URL")

        m3u8_url = stream_result.stdout.strip()
        safe_title = "".join(c if c.isalnum() or c in " -_" else "_" for c in title)
        output_file = os.path.join(output_dir, f"{safe_title}.m4a")

        process = subprocess.Popen(
            ["ffmpeg", "-y", "-i", m3u8_url, "-c", "copy", output_file],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE,
        )
    else:
        output_file = None
        process = subprocess.Popen(
            [
                "yt-dlp", "-f", "bestaudio",
                "--cookies", cookies_path,
                "--add-header", "Referer:https://twitter.com/",
                "--no-playlist",
                "-o", f"{output_dir}/%(title)s.%(ext)s",
                url,
            ],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE,
        )

    recording_id = uuid.uuid4().hex[:8]
    _active_recordings[recording_id] = {
        "process": process,
        "pid": process.pid,
        "url": url,
        "output_dir": output_dir,
        "start_time": time.time(),
        "before_files": before,
        "is_live": is_live,
        "title": title,
        "output_file": output_file if is_live else None,
    }

    mode_str = "Live recording from current moment" if is_live else "Downloading full recording"
    return {
        "success": True,
        "recording_id": recording_id,
        "mode": "live" if is_live else "recording",
        "title": title,
        "url": url,
        "output_dir": output_dir,
        "pid": process.pid,
        "message": f"{mode_str} (ID: {recording_id}). Use spaces_check to check progress, spaces_stop to stop.",
    }


def _spaces_check(arguments: dict) -> dict:
    """Check download/recording status."""
    import glob as glob_module

    recording_id = arguments.get("recording_id")
    if not recording_id:
        raise ValueError("Missing required parameter: recording_id")

    if recording_id not in _active_recordings:
        raise ValueError(f"No active download found with ID: {recording_id}")

    rec = _active_recordings[recording_id]
    process = rec["process"]
    output_dir = rec["output_dir"]
    before = rec["before_files"]
    elapsed = time.time() - rec["start_time"]

    poll = process.poll()

    output_file = rec.get("output_file")
    if not output_file:
        after = set(glob_module.glob(os.path.join(output_dir, "*")))
        new_files = after - before
        output_file = max(new_files, key=os.path.getmtime) if new_files else None

    file_info = {}
    if output_file and os.path.exists(output_file):
        file_info = {
            "path": output_file,
            "filename": os.path.basename(output_file),
            "size_mb": round(os.path.getsize(output_file) / (1024 * 1024), 2),
        }

    if poll is None:
        return {
            "recording_id": recording_id,
            "status": "downloading",
            "elapsed_seconds": round(elapsed),
            "elapsed_formatted": f"{int(elapsed // 60)}m {int(elapsed % 60)}s",
            "file": file_info,
            "message": f"Still downloading ({int(elapsed // 60)}m {int(elapsed % 60)}s elapsed)",
        }

    if poll == 0:
        del _active_recordings[recording_id]
        return {
            "recording_id": recording_id,
            "status": "complete",
            "elapsed_seconds": round(elapsed),
            "elapsed_formatted": f"{int(elapsed // 60)}m {int(elapsed % 60)}s",
            "file": file_info,
            "message": f"Download complete. Saved to {output_file}" if output_file else "Download complete",
        }

    stderr = process.stderr.read().decode() if process.stderr else ""
    del _active_recordings[recording_id]
    return {
        "recording_id": recording_id,
        "status": "error",
        "error": stderr.strip()[:500] or "Download failed",
        "message": "Download failed",
    }


def _spaces_stop(arguments: dict) -> dict:
    """Stop a live recording."""
    import glob as glob_module

    recording_id = arguments.get("recording_id")
    if not recording_id:
        raise ValueError("Missing required parameter: recording_id")

    if recording_id not in _active_recordings:
        raise ValueError(f"No active download found with ID: {recording_id}")

    rec = _active_recordings[recording_id]
    process = rec["process"]
    output_dir = rec["output_dir"]
    before = rec["before_files"]

    if process.poll() is None:
        process.send_signal(signal.SIGINT)
        try:
            process.wait(timeout=30)
        except Exception:
            process.terminate()
            try:
                process.wait(timeout=10)
            except Exception:
                process.kill()

    elapsed = time.time() - rec["start_time"]

    output_file = rec.get("output_file")
    if not output_file:
        after = set(glob_module.glob(os.path.join(output_dir, "*")))
        new_files = after - before
        output_file = max(new_files, key=os.path.getmtime) if new_files else None

    file_info = {}
    if output_file and os.path.exists(output_file):
        file_info = {
            "path": output_file,
            "filename": os.path.basename(output_file),
            "size_mb": round(os.path.getsize(output_file) / (1024 * 1024), 2),
        }

    del _active_recordings[recording_id]

    return {
        "success": True,
        "recording_id": recording_id,
        "status": "stopped",
        "elapsed_seconds": round(elapsed),
        "elapsed_formatted": f"{int(elapsed // 60)}m {int(elapsed % 60)}s",
        "file": file_info,
        "message": f"Stopped after {int(elapsed // 60)}m {int(elapsed % 60)}s. Saved to {output_file}" if output_file else "Stopped",
    }


def handle_request(request: dict) -> None:
    """Route JSON-RPC request to appropriate handler."""
    method = request.get("method")
    id = request.get("id")
    params = request.get("params", {})

    if method == "initialize":
        handle_initialize(id, params)
    elif method == "notifications/initialized":
        pass  # No response needed for notifications
    elif method == "tools/list":
        handle_tools_list(id)
    elif method == "tools/call":
        handle_tools_call(id, params)
    else:
        if id is not None:
            send_error(id, -32601, f"Method not found: {method}")


def main() -> None:
    """Main MCP server loop."""
    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue

        try:
            request = json.loads(line)
            handle_request(request)
        except json.JSONDecodeError:
            send_error(None, -32700, "Parse error")
        except Exception as e:
            send_error(None, -32603, f"Internal error: {str(e)}")


if __name__ == "__main__":
    main()
